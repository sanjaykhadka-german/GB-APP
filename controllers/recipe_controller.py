from flask import Blueprint, render_template, request, redirect, send_file, url_for, flash, jsonify
from sqlalchemy.sql import text
from decimal import Decimal
import sqlalchemy.exc
from datetime import datetime, timedelta
import pandas as pd
from io import BytesIO
from database import db
from models import Production, RecipeMaster, UsageReport, RawMaterialReport, ItemMaster
from models.usage_report import UsageReport
from models.recipe_master import RecipeMaster
from models.production import Production
# from models.joining import Joining  # REMOVED - joining table deprecated
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# Create a Blueprint for recipe routes
recipe_bp = Blueprint('recipe', __name__, template_folder='templates')

def get_monday_date(date_str):
    """Convert any date to the previous Monday if it's not already a Monday.
    Supports both YYYY-MM-DD and DD/MM/YYYY formats."""
    try:
        # Try YYYY-MM-DD format first
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        try:
            # Try DD/MM/YYYY format
            date = datetime.strptime(date_str, '%d/%m/%Y').date()
        except ValueError:
            raise ValueError("Date must be in YYYY-MM-DD or DD/MM/YYYY format")
    
    days_since_monday = date.weekday()  # Monday = 0, Sunday = 6
    monday = date - timedelta(days=days_since_monday)
    return monday

@recipe_bp.route('/recipe', methods=['GET', 'POST'])
def recipe_page():
    if request.method == 'POST':
        try:
            data = request.get_json() 
            recipes_data = data.get('recipes', [])

            if not recipes_data:
                return jsonify({'error': 'No recipes data provided.'}), 400
            
            # Validate descriptions are unique
            descriptions = {recipe.get('description') for recipe in recipes_data}
            if len(descriptions) > 1:
                return jsonify({'error': 'All recipes must have the same description.'}), 400

            # Process all recipes
            for recipe_data in recipes_data:
                recipe_id = recipe_data.get('recipe_id')
                recipe_code = recipe_data.get('recipe_code')
                description = recipe_data.get('description')
                finished_good_id = recipe_data.get('finished_good_id')
                raw_material_id = recipe_data.get('raw_material_id')
                kg_per_batch = recipe_data.get('kg_per_batch')

                if not all([recipe_code, description, finished_good_id, raw_material_id, kg_per_batch]):
                    return jsonify({'error': 'Required fields are missing.'}), 400
                
                # Validate kg_per_batch is a number
                try:
                    kg_per_batch = Decimal(kg_per_batch)
                    if kg_per_batch <= 0:
                        return jsonify({'error': 'Kg per batch cannot be negative or zero.'}), 400
                except (ValueError, TypeError):
                    return jsonify({'error': 'Invalid kg per batch value.'}), 400

                if recipe_id:  # Edit case
                    recipe = RecipeMaster.query.get_or_404(recipe_id)
                    recipe.recipe_code = recipe_code
                    recipe.description = description
                    recipe.finished_good_id = finished_good_id
                    recipe.raw_material_id = raw_material_id
                    recipe.kg_per_batch = kg_per_batch
                else:  # Add case
                    recipe = RecipeMaster(
                        recipe_code=recipe_code,
                        description=description,
                        finished_good_id=finished_good_id,
                        raw_material_id=raw_material_id,
                        kg_per_batch=kg_per_batch
                    )
                    db.session.add(recipe)

                db.session.flush()

            # Recalculate percentages for all recipes with the same recipe_code
            recipe_code = recipes_data[0]['recipe_code']
            recipes_to_update = RecipeMaster.query.filter(
                RecipeMaster.recipe_code == recipe_code
            ).all()

            total_quantity = sum(float(r.kg_per_batch) for r in recipes_to_update)
            for r in recipes_to_update:
                r.percentage = Decimal(round((float(r.kg_per_batch) / total_quantity) * 100, 2)) if total_quantity > 0 else Decimal('0.00')

            db.session.commit()
            return jsonify({'message': 'Recipes saved successfully!'}), 200

        except sqlalchemy.exc.IntegrityError as e:
            db.session.rollback()
            return jsonify({'error': 'Database error: Duplicate entry or invalid data.'}), 400
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'An error occurred: {str(e)}'}), 500

    # GET request: render the page
    search_recipe_code = request.args.get('recipe_code', '')
    search_description = request.args.get('description', '')
    edit_id = request.args.get('edit_id')
    
    # Get all items for dropdowns (any item can be used as component or assembly)
    all_items = ItemMaster.query.order_by(ItemMaster.item_code).all()

    return render_template('recipe/recipe.html', 
                         search_recipe_code=search_recipe_code,
                         search_description=search_description,
                         recipes=RecipeMaster.query.all(),
                         all_items=all_items,
                         current_page='recipe')

@recipe_bp.route('/recipe/delete/<int:id>', methods=['POST'])
def delete_recipe(id):
    try:
        recipe = RecipeMaster.query.get_or_404(id)
        recipe_code = recipe.recipe_code
        db.session.delete(recipe)
        db.session.commit()

        # Recalculate percentages for remaining recipes with the same recipe_code
        recipes_to_update = RecipeMaster.query.filter(RecipeMaster.recipe_code == recipe_code).all()
        if recipes_to_update:
            total_quantity = sum(float(r.kg_per_batch) for r in recipes_to_update)
            for r in recipes_to_update:
                r.percentage = Decimal(round((float(r.kg_per_batch) / total_quantity) * 100, 2)) if total_quantity > 0 else Decimal('0.00')
            db.session.commit()

        return jsonify({'message': 'Recipe deleted successfully!'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500

@recipe_bp.route('/autocomplete_recipe', methods=['GET'])
def autocomplete_recipe():
    search = request.args.get('query', '').strip()
    if not search:
        return jsonify([])
    try:
        query = text("SELECT recipe_code, description FROM recipe_master WHERE recipe_code LIKE :search LIMIT 10")
        results = db.session.execute(query, {"search": f"{search}%"}).fetchall()
        suggestions = [{"recipe_code": row[0], "description": row[1]} for row in results]
        return jsonify(suggestions)
    except Exception as e:
        print(f"Error fetching recipe autocomplete suggestions: {e}")
        return jsonify([])

@recipe_bp.route('/get_search_recipes', methods=['GET'])
def get_search_recipes():
    search_recipe_code = request.args.get('recipe_code', '').strip()
    search_description = request.args.get('description', '').strip()
    
    # Create aliases for the two ItemMaster joins
    from sqlalchemy.orm import aliased
    RawMaterialItem = aliased(ItemMaster)
    FinishedGoodItem = aliased(ItemMaster)
    
    # Join with both raw material and finished good items
    recipes_query = db.session.query(
        RecipeMaster,
        RawMaterialItem.item_code.label('raw_material_code'),
        RawMaterialItem.description.label('raw_material'),
        FinishedGoodItem.item_code.label('finished_good_code'),
        FinishedGoodItem.description.label('finished_good')
    ).join(
        RawMaterialItem,
        RecipeMaster.raw_material_id == RawMaterialItem.id
    ).join(
        FinishedGoodItem,
        RecipeMaster.finished_good_id == FinishedGoodItem.id
    )
    
    if search_recipe_code:
        recipes_query = recipes_query.filter(RecipeMaster.recipe_code.ilike(f"%{search_recipe_code}%"))
    if search_description:
        recipes_query = recipes_query.filter(RecipeMaster.description.ilike(f"%{search_description}%"))
    
    recipes = recipes_query.all()
    
    recipes_data = []
    for recipe in recipes:
        recipes_data.append({
            "id": recipe.RecipeMaster.id,
            "recipe_code": recipe.RecipeMaster.recipe_code,
            "description": recipe.RecipeMaster.description,
            "raw_material_code": recipe.raw_material_code,
            "raw_material": recipe.raw_material,
            "raw_material_id": recipe.RecipeMaster.raw_material_id,
            "finished_good_code": recipe.finished_good_code,
            "finished_good": recipe.finished_good,
            "finished_good_id": recipe.RecipeMaster.finished_good_id,
            "kg_per_batch": float(recipe.RecipeMaster.kg_per_batch) if recipe.RecipeMaster.kg_per_batch else 0.00,
            "percentage": float(recipe.RecipeMaster.percentage) if recipe.RecipeMaster.percentage else 0.00,
            "quantity_uom_id": recipe.RecipeMaster.quantity_uom_id
        })
    
    return jsonify(recipes_data)

@recipe_bp.route('/usage')
def usage():
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    # Query to get production and recipe usage data
    query = db.session.query(
        Production,
        RecipeMaster,
        ItemMaster.description.label('component_name')
    ).join(
        RecipeMaster,
        Production.production_code == RecipeMaster.recipe_code  # Join Production to RecipeMaster
    ).join(
        ItemMaster,
        RecipeMaster.raw_material_id == ItemMaster.id  # Join RecipeMaster to ItemMaster for raw material
    )
    
    # Apply date filters if provided
    if from_date and to_date:
        query = query.filter(
            Production.production_date >= from_date,
            Production.production_date <= to_date
        )
    
    # Get the results
    usage_data = query.all()
    
    # Group data by production date
    grouped_usage_data = {}
    for production, recipe, component_name in usage_data:
        date = production.production_date  # production_date is already a date object
        # Calculate the Monday of the week for the production_date
        week_commencing = get_monday_date(date.strftime('%Y-%m-%d'))
        
        if date not in grouped_usage_data:
            grouped_usage_data[date] = []
            
        grouped_usage_data[date].append({
            'week_commencing': week_commencing.strftime('%Y-%m-%d'),
            'production_date': production.production_date.strftime('%Y-%m-%d'),
            'production_code': production.production_code,
            'recipe_code': recipe.recipe_code,
            'component_material': component_name,
            'usage_kg': recipe.kg_per_batch * production.batches,  # Scale by batches
            'kg_per_batch': recipe.kg_per_batch,
            'percentage': recipe.percentage if recipe.percentage else 0.0
        })
    
    return render_template('recipe/usage.html',
                         grouped_usage_data=grouped_usage_data,
                         from_date=from_date,
                         to_date=to_date,
                         current_page='usage')

@recipe_bp.route('/usage/download')
def usage_download():
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    # Query to get usage data
    query = db.session.query(
        RecipeMaster,
        ItemMaster.description.label('component_name')
    ).join(
        ItemMaster,
        RecipeMaster.raw_material_id == ItemMaster.id
    )
    
    # Apply date filters if provided
    if from_date and to_date:
        query = query.filter(
            RecipeMaster.created_at >= from_date,
            RecipeMaster.created_at <= to_date
        )
    
    # Get the results
    usage_data = query.all()
    
    # Create Excel file
    data = []
    for recipe, component_name in usage_data:
        # Calculate the Monday of the week for the created_at date
        week_commencing = get_monday_date(recipe.created_at.date().strftime('%Y-%m-%d'))
        data.append({
            'Week Commencing': week_commencing.strftime('%Y-%m-%d'),
            'Production Date': recipe.created_at.strftime('%Y-%m-%d'),
            'Recipe Code': recipe.recipe_code,
            'Component Material': component_name,
            'Kg per Batch': recipe.kg_per_batch
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Usage Report', index=False)
    
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'usage_report_{from_date}_{to_date}.xlsx' if from_date and to_date else 'usage_report.xlsx'
    )   

@recipe_bp.route('/recipe/upload-excel', methods=['POST'])
def upload_recipe_excel():
    try:
        if 'excel_file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Please upload a valid Excel file (.xlsx or .xls)'}), 400
        
        # Read the Excel file
        import openpyxl
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        # Expected columns: Recipe Code, Description, Finished Good Code, Raw Material Code, Kg Per Batch, UOM, Is Active
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # Validate required headers
        required_headers = ['Recipe Code', 'Description', 'Finished Good Code', 'Raw Material Code', 'Kg Per Batch']
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            return jsonify({'error': f'Missing required columns: {", ".join(missing_headers)}'}), 400
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Process each row (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row):  # Skip empty rows
                    continue
                
                # Create dictionary from row data
                row_data = dict(zip(headers, row))
                
                # Skip if required fields are empty
                if not all([row_data.get('Recipe Code'), row_data.get('Description'), 
                           row_data.get('Finished Good Code'), row_data.get('Raw Material Code'), 
                           row_data.get('Kg Per Batch')]):
                    error_count += 1
                    errors.append(f'Row {row_num}: Missing required fields')
                    continue
                
                recipe_code = str(row_data['Recipe Code']).strip()
                description = str(row_data['Description']).strip()
                finished_good_code = str(row_data['Finished Good Code']).strip()
                raw_material_code = str(row_data['Raw Material Code']).strip()
                
                # Validate kg_per_batch is a number
                try:
                    kg_per_batch = float(row_data['Kg Per Batch'])
                    if kg_per_batch <= 0:
                        error_count += 1
                        errors.append(f'Row {row_num}: Kg per batch must be greater than 0')
                        continue
                except (ValueError, TypeError):
                    error_count += 1
                    errors.append(f'Row {row_num}: Invalid kg per batch value')
                    continue
                
                # Find finished good item
                finished_good = ItemMaster.query.filter_by(item_code=finished_good_code).first()
                if not finished_good:
                    error_count += 1
                    errors.append(f'Row {row_num}: Finished good code "{finished_good_code}" not found')
                    continue
                
                # Find raw material item
                raw_material = ItemMaster.query.filter_by(item_code=raw_material_code).first()
                if not raw_material:
                    error_count += 1
                    errors.append(f'Row {row_num}: Raw material code "{raw_material_code}" not found')
                    continue
                
                # Check if recipe already exists
                existing_recipe = RecipeMaster.query.filter(
                    RecipeMaster.finished_good_id == finished_good.id,
                    RecipeMaster.raw_material_id == raw_material.id
                ).first()
                if existing_recipe:
                    error_count += 1
                    errors.append(f'Row {row_num}: Recipe for {finished_good_code} -> {raw_material_code} already exists')
                    continue
                
                # Find UOM if specified
                uom_id = None
                if row_data.get('UOM'):
                    from models.uom import UOM
                    uom = UOM.query.filter_by(UOMName=str(row_data['UOM']).strip()).first()
                    if uom:
                        uom_id = uom.UOMID
                
                # Create new recipe
                recipe = RecipeMaster()
                recipe.recipe_code = recipe_code
                recipe.description = description
                recipe.finished_good_id = finished_good.id
                recipe.raw_material_id = raw_material.id
                recipe.kg_per_batch = kg_per_batch
                recipe.quantity_uom_id = uom_id
                
                # Set active status
                if row_data.get('Is Active'):
                    value = str(row_data['Is Active']).strip().lower()
                    recipe.is_active = value in ['true', '1', 'yes', 'y']
                else:
                    recipe.is_active = True  # Default to active
                
                db.session.add(recipe)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f'Row {row_num}: {str(e)}')
        
        # Calculate percentages for all recipes after upload
        if success_count > 0:
            db.session.flush()  # Ensure all recipes are in the database
            
            # Get all unique recipe codes that were uploaded
            uploaded_recipe_codes = set()
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if any(row):
                    row_data = dict(zip(headers, row))
                    if row_data.get('Recipe Code'):
                        uploaded_recipe_codes.add(str(row_data['Recipe Code']).strip())
            
            # Recalculate percentages for each recipe group
            for recipe_code in uploaded_recipe_codes:
                recipes_in_group = RecipeMaster.query.filter_by(recipe_code=recipe_code).all()
                total_kg = sum(float(r.kg_per_batch) for r in recipes_in_group)
                
                if total_kg > 0:
                    for recipe in recipes_in_group:
                        recipe.percentage = round((float(recipe.kg_per_batch) / total_kg) * 100, 2)
        
        db.session.commit()
        
        message = f'Upload completed: {success_count} recipes added successfully'
        if error_count > 0:
            message += f', {error_count} errors occurred'
            if len(errors) <= 5:  # Show first 5 errors
                message += f'. Errors: {"; ".join(errors)}'
        
        return jsonify({'message': message}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process Excel file: {str(e)}'}), 500

@recipe_bp.route('/recipe/download-excel', methods=['GET'])
def download_recipe_excel():
    try:
        # Get search parameters
        search_recipe_code = request.args.get('recipe_code', '').strip()
        search_description = request.args.get('description', '').strip()
        
        # Build query with same logic as get_search_recipes
        from sqlalchemy.orm import aliased
        RawMaterialItem = aliased(ItemMaster)
        FinishedGoodItem = aliased(ItemMaster)
        
        query = db.session.query(
            RecipeMaster,
            RawMaterialItem.item_code.label('raw_material_code'),
            RawMaterialItem.description.label('raw_material_name'),
            FinishedGoodItem.item_code.label('finished_good_code'),
            FinishedGoodItem.description.label('finished_good_name')
        ).join(
            RawMaterialItem,
            RecipeMaster.raw_material_id == RawMaterialItem.id
        ).join(
            FinishedGoodItem,
            RecipeMaster.finished_good_id == FinishedGoodItem.id
        )
        
        if search_recipe_code:
            query = query.filter(RecipeMaster.recipe_code.ilike(f"%{search_recipe_code}%"))
        if search_description:
            query = query.filter(RecipeMaster.description.ilike(f"%{search_description}%"))
        
        recipes = query.all()
        
        # Create workbook
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Recipe Master"
        
        # Define headers
        headers = [
            'Recipe Code', 'Description', 'Finished Good Code', 'Finished Good Name',
            'Raw Material Code', 'Raw Material Name', 'Kg Per Batch', 'Percentage', 
            'UOM', 'Is Active'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row, recipe_data in enumerate(recipes, 2):
            recipe = recipe_data.RecipeMaster
            sheet.cell(row=row, column=1, value=recipe.recipe_code)
            sheet.cell(row=row, column=2, value=recipe.description)
            sheet.cell(row=row, column=3, value=recipe_data.finished_good_code)
            sheet.cell(row=row, column=4, value=recipe_data.finished_good_name)
            sheet.cell(row=row, column=5, value=recipe_data.raw_material_code)
            sheet.cell(row=row, column=6, value=recipe_data.raw_material_name)
            sheet.cell(row=row, column=7, value=recipe.kg_per_batch)
            sheet.cell(row=row, column=8, value=recipe.percentage)
            
            # UOM
            uom_name = ''
            if recipe.quantity_uom_id:
                from models.uom import UOM
                uom = UOM.query.get(recipe.quantity_uom_id)
                if uom:
                    uom_name = uom.UOMName
            sheet.cell(row=row, column=9, value=uom_name)
            
            sheet.cell(row=row, column=10, value='Yes' if recipe.is_active else 'No')
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        import io
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f'recipe_master_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate Excel file: {str(e)}'}), 500

@recipe_bp.route('/recipe/download-template', methods=['GET'])
def download_recipe_template():
    try:
        # Create workbook with template structure
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Recipe Master Template"
        
        # Define headers
        headers = [
            'Recipe Code', 'Description', 'Finished Good Code', 'Raw Material Code',
            'Kg Per Batch', 'UOM', 'Is Active'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add sample data row
        sample_data = [
            'REC001', 'Sample Recipe Description', 'FG001', 'RM001', '10.5', 'KG', 'Yes'
        ]
        
        for col, value in enumerate(sample_data, 1):
            sheet.cell(row=2, column=col, value=value)
        
        # Add instructions in a separate sheet
        instructions_sheet = workbook.create_sheet("Instructions")
        instructions = [
            "RECIPE MASTER UPLOAD INSTRUCTIONS",
            "",
            "Required Columns:",
            "- Recipe Code: Unique identifier for the recipe group",
            "- Description: Recipe description",
            "- Finished Good Code: Item code of what is being made (must exist in Item Master)",
            "- Raw Material Code: Item code of component needed (must exist in Item Master)",
            "- Kg Per Batch: Quantity of component needed per batch (must be > 0)",
            "",
            "Optional Columns:",
            "- UOM: Unit of measure (must match existing UOMs)",
            "- Is Active: Yes/No (defaults to Yes if not specified)",
            "",
            "Notes:",
            "- Do not modify the header row",
            "- Empty rows will be skipped",
            "- Item codes must exist in the Item Master",
            "- Duplicate finished good + raw material combinations will be rejected",
            "- Percentages will be calculated automatically within each recipe group",
            "- Boolean fields accept: Yes/No, True/False, 1/0, Y/N",
            "",
            "Example:",
            "Recipe Code: FRANK001",
            "Description: Frankfurter Recipe",
            "Finished Good Code: 2006 (Frankfurter WIP)",
            "Raw Material Code: RM001 (Pork 80CL)",
            "Kg Per Batch: 40.5"
        ]
        
        for row, instruction in enumerate(instructions, 1):
            cell = instructions_sheet.cell(row=row, column=1, value=instruction)
            if row == 1:  # Title
                cell.font = Font(bold=True, size=14)
            elif instruction.endswith(":"):  # Section headers
                cell.font = Font(bold=True)
        
        # Auto-adjust column widths for both sheets
        for sheet_obj in [sheet, instructions_sheet]:
            for column in sheet_obj.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                sheet_obj.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        import io
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name='recipe_master_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate template: {str(e)}'}), 500

@recipe_bp.route('/recipe/upload', methods=['GET', 'POST'])
def recipe_upload():
    """Render the upload page for Recipe Master Excel upload and handle file uploads."""
    if request.method == 'POST':
        try:
            if 'file' not in request.files:
                flash('No file uploaded', 'error')
                return render_template('recipe/upload.html', current_page='recipe')
            
            file = request.files['file']
            if file.filename == '':
                flash('No file selected', 'error')
                return render_template('recipe/upload.html', current_page='recipe')
            
            if not file.filename.lower().endswith(('.xlsx', '.xls')):
                flash('Please upload a valid Excel file (.xlsx or .xls)', 'error')
                return render_template('recipe/upload.html', current_page='recipe')
            
            # Get sheet name from form
            sheet_name = request.form.get('sheet_name', '').strip() or 'Recipe Master'
            
            # Process the file using existing upload logic
            import openpyxl
            workbook = openpyxl.load_workbook(file)
            
            # Try to get the specified sheet
            try:
                sheet = workbook[sheet_name]
            except KeyError:
                # If specified sheet doesn't exist, try active sheet
                sheet = workbook.active
                flash(f'Sheet "{sheet_name}" not found, using active sheet instead', 'warning')
            
            # Expected columns
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Validate required headers
            required_headers = ['Recipe Code', 'Description', 'Finished Good Code', 'Raw Material Code', 'Kg Per Batch']
            missing_headers = [h for h in required_headers if h not in headers]
            if missing_headers:
                flash(f'Missing required columns: {", ".join(missing_headers)}', 'error')
                return render_template('recipe/upload.html', current_page='recipe')
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Process each row (skip header)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not any(row):  # Skip empty rows
                        continue
                    
                    # Create dictionary from row data
                    row_data = dict(zip(headers, row))
                    
                    # Skip if required fields are empty
                    if not all([row_data.get('Recipe Code'), row_data.get('Description'), 
                               row_data.get('Finished Good Code'), row_data.get('Raw Material Code'), 
                               row_data.get('Kg Per Batch')]):
                        error_count += 1
                        errors.append(f'Row {row_num}: Missing required fields')
                        continue
                    
                    recipe_code = str(row_data['Recipe Code']).strip()
                    description = str(row_data['Description']).strip()
                    finished_good_code = str(row_data['Finished Good Code']).strip()
                    raw_material_code = str(row_data['Raw Material Code']).strip()
                    
                    # Validate kg_per_batch is a number
                    try:
                        kg_per_batch = float(row_data['Kg Per Batch'])
                        if kg_per_batch <= 0:
                            error_count += 1
                            errors.append(f'Row {row_num}: Kg per batch must be greater than 0')
                            continue
                    except (ValueError, TypeError):
                        error_count += 1
                        errors.append(f'Row {row_num}: Invalid kg per batch value')
                        continue
                    
                    # Find finished good item
                    finished_good = ItemMaster.query.filter_by(item_code=finished_good_code).first()
                    if not finished_good:
                        error_count += 1
                        errors.append(f'Row {row_num}: Finished good code "{finished_good_code}" not found')
                        continue
                    
                    # Find raw material item
                    raw_material = ItemMaster.query.filter_by(item_code=raw_material_code).first()
                    if not raw_material:
                        error_count += 1
                        errors.append(f'Row {row_num}: Raw material code "{raw_material_code}" not found')
                        continue
                    
                    # Check if recipe already exists
                    existing_recipe = RecipeMaster.query.filter(
                        RecipeMaster.finished_good_id == finished_good.id,
                        RecipeMaster.raw_material_id == raw_material.id
                    ).first()
                    if existing_recipe:
                        error_count += 1
                        errors.append(f'Row {row_num}: Recipe for {finished_good_code} -> {raw_material_code} already exists')
                        continue
                    
                    # Find UOM if specified
                    uom_id = None
                    if row_data.get('UOM'):
                        from models.uom import UOM
                        uom = UOM.query.filter_by(UOMName=str(row_data['UOM']).strip()).first()
                        if uom:
                            uom_id = uom.UOMID
                    
                    # Create new recipe
                    recipe = RecipeMaster()
                    recipe.recipe_code = recipe_code
                    recipe.description = description
                    recipe.finished_good_id = finished_good.id
                    recipe.raw_material_id = raw_material.id
                    recipe.kg_per_batch = kg_per_batch
                    recipe.quantity_uom_id = uom_id
                    
                    # Set active status
                    if row_data.get('Is Active'):
                        value = str(row_data['Is Active']).strip().lower()
                        recipe.is_active = value in ['true', '1', 'yes', 'y']
                    else:
                        recipe.is_active = True  # Default to active
                    
                    db.session.add(recipe)
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'Row {row_num}: {str(e)}')
            
            # Calculate percentages for all recipes after upload
            if success_count > 0:
                db.session.flush()  # Ensure all recipes are in the database
                
                # Get all unique recipe codes that were uploaded
                uploaded_recipe_codes = set()
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if any(row):
                        row_data = dict(zip(headers, row))
                        if row_data.get('Recipe Code'):
                            uploaded_recipe_codes.add(str(row_data['Recipe Code']).strip())
                
                # Recalculate percentages for each recipe group
                for recipe_code in uploaded_recipe_codes:
                    recipes_in_group = RecipeMaster.query.filter_by(recipe_code=recipe_code).all()
                    total_kg = sum(float(r.kg_per_batch) for r in recipes_in_group)
                    
                    if total_kg > 0:
                        for recipe in recipes_in_group:
                            recipe.percentage = round((float(recipe.kg_per_batch) / total_kg) * 100, 2)
            
            db.session.commit()
            
            message = f'Upload completed: {success_count} recipes added successfully'
            if error_count > 0:
                message += f', {error_count} errors occurred'
                if len(errors) <= 5:  # Show first 5 errors
                    message += f'. Errors: {"; ".join(errors)}'
                flash(message, 'warning')
            else:
                flash(message, 'success')
            
            return render_template('recipe/upload.html', current_page='recipe')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Failed to process file: {str(e)}', 'error')
            return render_template('recipe/upload.html', current_page='recipe')
    
    # GET request - just render the upload page
    return render_template('recipe/upload.html', current_page='recipe')

@recipe_bp.route('/raw_material_report', methods=['GET'])
def raw_material_report():
    try:
        # Get week commencing filter from request
        week_commencing = request.args.get('week_commencing')
        
        # Base query for weekly data - using current schema
        raw_material_query = """
        SELECT 
            DATE(p.production_date - INTERVAL (WEEKDAY(p.production_date)) DAY) as week_commencing,
            im.description as component_material,
            im.id as component_item_id,
            SUM(p.total_kg * r.percentage / 100) as total_usage
        FROM production p
        JOIN recipe_master r ON p.production_code = r.recipe_code
        JOIN item_master im ON r.raw_material_id = im.id
        """
        
        # Add date filter to the query
        params = {}
        if week_commencing:
            raw_material_query += """ 
            WHERE DATE(p.production_date - INTERVAL (WEEKDAY(p.production_date)) DAY) = :week_commencing
            """
            params['week_commencing'] = datetime.strptime(week_commencing, '%Y-%m-%d').date()
        
        raw_material_query += """
        GROUP BY 
            DATE(p.production_date - INTERVAL (WEEKDAY(p.production_date)) DAY),
            im.description,
            im.id
        ORDER BY week_commencing DESC, im.description
        """
        
        results = db.session.execute(text(raw_material_query), params).fetchall()
        
        # Clear existing records for the week
        if week_commencing:
            delete_query = "DELETE FROM raw_material_report WHERE week_commencing = :week_commencing"
            delete_params = {'week_commencing': datetime.strptime(week_commencing, '%Y-%m-%d').date()}
            db.session.execute(text(delete_query), delete_params)
        
        # Save results to raw_material_report table
        for result in results:
            report = RawMaterialReport(
                production_date=result.week_commencing,  # Using week_commencing as production_date
                week_commencing=result.week_commencing,
                raw_material=result.component_material,
                raw_material_id=result.component_item_id,
                meat_required=float(result.total_usage),
                created_at=datetime.now()
            )
            db.session.add(report)
        
        db.session.commit()
        
        # Convert to list of dictionaries for template
        raw_material_data = [
            {
                'week_commencing': result.week_commencing.strftime('%d/%m/%Y'),
                'raw_material': result.component_material,
                'usage': round(float(result.total_usage), 2)
            }
            for result in results
        ]
        
        return render_template('recipe/raw_material_report.html', 
                             raw_material_data=raw_material_data,
                             week_commencing=week_commencing,
                             current_page='raw_material_report')
        
    except Exception as e:
        flash(f"An error occurred: {str(e)}", 'error')
        return render_template('recipe/raw_material_report.html', 
                             raw_material_data=[],
                             week_commencing=week_commencing,
                             current_page='raw_material_report')

@recipe_bp.route('/raw_material_download', methods=['GET'])
def raw_material_download():
    try:
        week_commencing = request.args.get('week_commencing')
        
        # Base query for weekly data - using current schema
        raw_material_query = """
        SELECT 
            DATE(p.production_date - INTERVAL (WEEKDAY(p.production_date)) DAY) as week_commencing,
            im.description as component_material,
            im.id as component_item_id,
            SUM(p.total_kg * r.percentage / 100) as total_usage
        FROM production p
        JOIN recipe_master r ON p.production_code = r.recipe_code
        JOIN item_master im ON r.raw_material_id = im.id
        """
        
        # Add date filter to the query
        params = {}
        if week_commencing:
            raw_material_query += """ 
            WHERE DATE(p.production_date - INTERVAL (WEEKDAY(p.production_date)) DAY) = :week_commencing
            """
            params['week_commencing'] = datetime.strptime(week_commencing, '%Y-%m-%d').date()
        
        raw_material_query += """
        GROUP BY 
            DATE(p.production_date - INTERVAL (WEEKDAY(p.production_date)) DAY),
            im.description,
            im.id
        ORDER BY week_commencing DESC, im.description
        """
        
        results = db.session.execute(text(raw_material_query), params).fetchall()
        
        # Convert to list of dictionaries for Excel
        data = [
            {
                'Week Commencing': result.week_commencing.strftime('%d/%m/%Y'),
                'Component Material': result.component_material,
                'Total Usage (kg)': round(float(result.total_usage), 2)
            }
            for result in results
        ]
        
        if not data:
            flash("No data available for the selected week.", 'warning')
            return redirect(url_for('recipe.raw_material_report'))
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel file
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Raw Material Report', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Raw Material Report']
            
            # Add formatting
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#D7E4BC',
                'border': 1
            })
            
            # Write the column headers with the defined format
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            # Adjust column widths
            worksheet.set_column('A:A', 15)  # Week Commencing
            worksheet.set_column('B:B', 30)  # Component Material
            worksheet.set_column('C:C', 15)  # Total Usage
        
        output.seek(0)
        
        filename = f'raw_material_report_{week_commencing}.xlsx' if week_commencing else 'raw_material_report.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f"Error generating Excel file: {str(e)}", 'error')
        return redirect(url_for('recipe.raw_material_report')) 