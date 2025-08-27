from flask import Blueprint, render_template, request, jsonify, send_file, flash, session, redirect, url_for
from database import db
from models.item_master import ItemMaster, ItemAllergen
from models.category import Category
from models.department import Department
from models.machinery import Machinery
from models.uom import UOM
from models.allergen import Allergen
from models.item_type import ItemType
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import os
from sqlalchemy.orm import joinedload
from sqlalchemy import asc
import logging
import pytz

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

item_master_bp = Blueprint('item_master', __name__)

@item_master_bp.route('/item-master', methods=['GET'])
def item_master_list():
    # Preload related data for efficiency
    items = ItemMaster.query.options(
        joinedload(ItemMaster.category),
        joinedload(ItemMaster.department),
        joinedload(ItemMaster.uom),
        joinedload(ItemMaster.item_type)
    ).order_by(asc(ItemMaster.item_code)).all()
    
    categories = Category.query.all()
    departments = Department.query.all()
    item_types = ItemType.query.all()
    
    return render_template(
        'item_master/list.html', 
        items=items,
        categories=categories,
        departments=departments,
        item_types=item_types,
        current_page='item_master'
    )

@item_master_bp.route('/item-master/create', methods=['GET'])
def item_master_create():
    categories = Category.query.all()
    departments = Department.query.all()
    machineries = Machinery.query.all()
    uoms = UOM.query.all()
    allergens = Allergen.query.all()
    item_types = ItemType.query.all()
    wip_items = ItemMaster.query.join(ItemType).filter(ItemType.type_name == 'WIP').all()
    wipf_items = ItemMaster.query.join(ItemType).filter(ItemType.type_name == 'WIPF').all()
    
    return render_template('item_master/create.html', 
                           categories=categories, 
                           departments=departments, 
                           machineries=machineries, 
                           uoms=uoms, 
                           allergens=allergens,
                           item_types=item_types,
                           wip_items=wip_items,
                           wipf_items=wipf_items,
                           current_page='item_master')

@item_master_bp.route('/item-master/edit/<int:id>', methods=['GET'])
def item_master_edit(id):
    # Get the item
    item = ItemMaster.query.get_or_404(id)
    
    # Get all lookup data
    categories = Category.query.all()
    departments = Department.query.all()
    machineries = Machinery.query.all()
    uoms = UOM.query.all()
    allergens = Allergen.query.all()
    item_types = ItemType.query.all()
    
    # Get WIP and WIPF items for component dropdowns
    wip_items = ItemMaster.query.join(ItemType).filter(ItemType.type_name == 'WIP').all()
    wipf_items = ItemMaster.query.join(ItemType).filter(ItemType.type_name == 'WIPF').all()

    return render_template('item_master/edit.html',
                         categories=categories,
                         departments=departments,
                         machineries=machineries,
                         uoms=uoms,
                         allergens=allergens,
                         item=item,
                         item_types=item_types,
                         wip_items=wip_items,
                         wipf_items=wipf_items,
                         current_page='item_master')

@item_master_bp.route('/item_master/delete/<int:id>', methods=['POST'])
def item_master_delete(id):
    item = ItemMaster.query.get_or_404(id)
    db.session.delete(item)
    db.session.commit()
    flash(f'Item {item.item_code} has been deleted.', 'success')
    return redirect(url_for('item_master.item_master_list'))

@item_master_bp.route('/item_master/get-item-info/<int:item_id>')
def get_item_info_json(item_id):
    try:
        item = ItemMaster.query.options(
            joinedload(ItemMaster.allergens)
        ).get(item_id)
        
        if not item:
            return jsonify({'error': 'Item not found'}), 404
            
        # Get allergen IDs associated with the item
        allergen_ids = [item_allergen.allergens_id for item_allergen in item.allergens]
        
        item_data = {
            'id': item.id,
            'item_code': item.item_code,
            'description': item.description,
            'category_id': item.category_id,
            'department_id': item.department_id,
            'uom_id': item.uom_id,
            'machinery_id': item.machinery_id,
            'min_level': item.min_level,
            'max_level': item.max_level,
            'price_per_kg': item.price_per_kg,
            'price_per_uom': item.price_per_uom,
            'supplier_name': item.supplier_name,
            'is_active': item.is_active,
            'is_make_to_order': item.is_make_to_order,
            'fw': item.fw,
            'loss_percentage': item.loss_percentage,
            'kg_per_unit': item.kg_per_unit,
            'units_per_bag': item.units_per_bag,
            'avg_weight_per_unit': item.avg_weight_per_unit,
            #'calculation_factor': item.calculation_factor,
            'item_type_id': item.item_type_id,
            'wip_item_id': item.wip_item_id,
            'wipf_item_id': item.wipf_item_id,
            'allergens': allergen_ids
        }
        return jsonify(item_data)
        
    except Exception as e:
        logger.error(f"Error fetching item info for ID {item_id}: {str(e)}")
        return jsonify({'error': 'Internal Server Error'}), 500

@item_master_bp.route('/item-master/create', methods=['POST'])
@item_master_bp.route('/item-master/edit/<int:id>', methods=['PUT'])
def save_item(id=None):
    try:
        data = request.get_json()
        logger.info(f"Received data for save/update: {data}")
        
        # Get item type
        item_type_name = data.get('item_type')
        item_type = ItemType.query.filter_by(type_name=item_type_name).first()
        if not item_type:
            return jsonify({'success': False, 'message': 'Item type is required'}), 400
            
        # Check if department and machinery are required
        requires_dept_mach = item_type_name in ['FG', 'WIPF', 'WIP']
        department_id = data.get('department_id')
        machinery_id = data.get('machinery_id')
        
        if requires_dept_mach:
            if not department_id:
                return jsonify({'success': False, 'message': f'Department is required for {item_type_name} items'}), 400
            if not machinery_id:
                return jsonify({'success': False, 'message': f'Machinery is required for {item_type_name} items'}), 400
        
        # Get or create item
        if id:
            item = ItemMaster.query.get_or_404(id)
            item.updated_at = datetime.now(pytz.timezone('Australia/Sydney'))
        else:
            item = ItemMaster()
            item.created_at = datetime.now(pytz.timezone('Australia/Sydney'))
            
        # Update basic fields
        item.item_code = data.get('item_code')
        item.description = data.get('description')
        item.item_type_id = item_type.id
        
        # Update department and machinery
        item.department_id = int(department_id) if department_id else None
        item.machinery_id = int(machinery_id) if machinery_id else None
        
        # Update other fields, handling None for empty strings
        item.category_id = int(data.get('category_id')) if data.get('category_id') else None
        item.uom_id = int(data.get('uom_id')) if data.get('uom_id') else None
        item.min_level = float(data.get('min_level')) if data.get('min_level') else None
        item.max_level = float(data.get('max_level')) if data.get('max_level') else None
        item.price_per_kg = float(data.get('price_per_kg')) if data.get('price_per_kg') else None
        item.price_per_uom = float(data.get('price_per_uom')) if data.get('price_per_uom') else None
        item.supplier_name = data.get('supplier_name') or None
        item.is_active = data.get('is_active', False)
        item.is_make_to_order = data.get('is_make_to_order', False)
        item.fw = data.get('fw', False)
        item.loss_percentage = float(data.get('loss_percentage')) if data.get('loss_percentage') else None
        item.kg_per_unit = float(data.get('kg_per_unit')) if data.get('kg_per_unit') else None
        item.units_per_bag = float(data.get('units_per_bag')) if data.get('units_per_bag') else None
        item.avg_weight_per_unit = float(data.get('avg_weight_per_unit')) if data.get('avg_weight_per_unit') else None
        #item.calculation_factor = float(data.get('calculation_factor')) if data.get('calculation_factor') else None
        
        # Update component relationships
        item.wip_item_id = int(data.get('wip_item_id')) if data.get('wip_item_id') else None
        item.wipf_item_id = int(data.get('wipf_item_id')) if data.get('wipf_item_id') else None
        
        # Update Allergens
        allergen_ids = data.get('allergen_ids', [])
        
        # Clear existing allergens
        ItemAllergen.query.filter_by(item_id=item.id).delete()
        
        # Add new allergens
        if allergen_ids:
            for allergen_id in allergen_ids:
                item_allergen = ItemAllergen(item_id=item.id, allergen_id=allergen_id)
                db.session.add(item_allergen)

        if not id:
            db.session.add(item)
            
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Item saved successfully', 'id': item.id})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error saving item: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@item_master_bp.route('/get_items', methods=['GET'])
def get_items():
    try:
        # Base query
        query = ItemMaster.query

        # Search filters
        search_code = request.args.get('item_code')
        search_code_or_desc = request.args.get('item_code_or_description')
        search_desc = request.args.get('description')
        search_type = request.args.get('item_type')
        search_category = request.args.get('category')
        search_department = request.args.get('department')
        search_is_active = request.args.get('is_active')

        # Sorting
        sort_by = request.args.get('sort_by', 'item_code')
        sort_order = request.args.get('sort_order', 'asc')

        # Determine which tables we need to join based on search and sort
        required_joins = set()
        if search_type or sort_by == 'item_type':
            required_joins.add('item_type')
        if search_category or sort_by == 'category':
            required_joins.add('category')
        if search_department or sort_by == 'department':
            required_joins.add('department')

        # Apply joins
        if 'item_type' in required_joins:
            query = query.join(ItemType, isouter=True)
        if 'category' in required_joins:
            query = query.join(Category, isouter=True)
        if 'department' in required_joins:
            query = query.join(Department, isouter=True)

        # Apply filters
        if search_code_or_desc:
            # Search in both item_code and description
            from sqlalchemy import or_
            query = query.filter(
                or_(
                    ItemMaster.item_code.ilike(f'%{search_code_or_desc}%'),
                    ItemMaster.description.ilike(f'%{search_code_or_desc}%')
                )
            )
        elif search_code:
            query = query.filter(ItemMaster.item_code.ilike(f'%{search_code}%'))
        if search_desc:
            query = query.filter(ItemMaster.description.ilike(f'%{search_desc}%'))
        if search_type:
            query = query.filter(ItemType.type_name == search_type)
        if search_category:
            query = query.filter(Category.name == search_category)
        if search_department:
            query = query.filter(Department.departmentName == search_department)
        if search_is_active is not None:
            query = query.filter(ItemMaster.is_active == (search_is_active.lower() in ['true', '1']))

        # Map API sort columns to model attributes
        sort_map = {
            'item_code': ItemMaster.item_code,
            'description': ItemMaster.description,
            'item_type': ItemType.type_name,
            'category': Category.name,
            'department': Department.departmentName,
            #'calculation_factor': ItemMaster.calculation_factor,
            'units_per_bag': ItemMaster.units_per_bag,
            'avg_weight_per_unit': ItemMaster.avg_weight_per_unit,
            'is_active': ItemMaster.is_active
        }
        
        sort_column = sort_map.get(sort_by, ItemMaster.item_code)

        # Apply sorting direction
        if sort_order == 'desc':
            query = query.order_by(sort_column.desc())
        else:
            query = query.order_by(sort_column.asc())

        # Eager load related objects to avoid N+1 query problem
        items = query.options(
            joinedload(ItemMaster.category),
            joinedload(ItemMaster.department),
            joinedload(ItemMaster.uom),
            joinedload(ItemMaster.item_type)
        ).all()
        
        # Format data for JSON response
        item_list = []
        for item in items:
            # Convert User objects to strings if needed
            created_by = str(item.created_by) if item.created_by else None
            updated_by = str(item.updated_by) if item.updated_by else None
            
            item_list.append({
                'id': item.id,
                'item_code': item.item_code,
                'description': item.description,
                'item_type': item.item_type.type_name if item.item_type else '',
                'category': item.category.name if item.category else '',
                'department': item.department.departmentName if item.department else '',
                'uom': item.uom.UOMName if item.uom else '',
                'min_level': item.min_level,
                'max_level': item.max_level,
                'price_per_kg': item.price_per_kg,
                'price_per_uom': item.price_per_uom,
                'supplier_name': item.supplier_name,
                # 'calculation_factor': item.calculation_factor,
                'units_per_bag': item.units_per_bag,
                'avg_weight_per_unit': item.avg_weight_per_unit,
                'is_active': item.is_active,
                'created_by': created_by,
                'created_at': item.created_at.strftime('%d-%m-%Y %H:%M') if item.created_at else '',
                'updated_by': updated_by,
                'updated_at': item.updated_at.strftime('%d-%m-%Y %H:%M') if item.updated_at else ''
            })
            
        return jsonify(item_list)
    except Exception as e:
        logger.error(f"Error in get_items: {str(e)}", exc_info=True)
        return jsonify({'error': 'An internal error occurred. Please check the server logs.'}), 500

@item_master_bp.route('/item-master/upload-excel', methods=['POST'])
def upload_excel():
    try:
        if 'excel_file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Please upload a valid Excel file (.xlsx or .xls)'}), 400
        
        # Read the Excel file
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        # Expected columns: Item Code, Description, Type, Category, Department, UOM, Min Level, Max Level, Price Per Kg, Is Make To Order, Kg Per Unit, Units Per Bag, Loss Percentage, Is Active
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # Validate required headers
        required_headers = ['Item Code', 'Description', 'Type']
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            return jsonify({'error': f'Missing required columns: {", ".join(missing_headers)}'}), 400
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Process each row (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row):  # Skip empty rows
                    continue
                
                # Create dictionary from row data
                row_data = dict(zip(headers, row))
                
                # Skip if required fields are empty
                if not row_data.get('Item Code') or not row_data.get('Description') or not row_data.get('Type'):
                    error_count += 1
                    errors.append(f'Row {row_num}: Missing required fields')
                    continue
                
                item_code = str(row_data['Item Code']).strip()
                description = str(row_data['Description']).strip()
                item_type = str(row_data['Type']).strip()
                
                # Check if item already exists
                existing_item = ItemMaster.query.filter_by(item_code=item_code).first()
                if existing_item:
                    error_count += 1
                    errors.append(f'Row {row_num}: Item code {item_code} already exists')
                    continue
                
                # Validate item type
                valid_item_type = ItemType.query.filter_by(type_name=item_type).first()
                if not valid_item_type:
                    error_count += 1
                    errors.append(f'Row {row_num}: Invalid item type "{item_type}"')
                    continue
                
                # Create new item
                item = ItemMaster()
                item.item_code = item_code
                item.description = description
                item.item_type = item_type
                
                # Optional fields
                if row_data.get('Category'):
                    category = Category.query.filter_by(name=str(row_data['Category']).strip()).first()
                    if category:
                        item.category_id = category.id
                
                if row_data.get('Department'):
                    department = Department.query.filter_by(departmentName=str(row_data['Department']).strip()).first()
                    if department:
                        item.department_id = department.id
                
                if row_data.get('UOM'):
                    uom = UOM.query.filter_by(UOMName=str(row_data['UOM']).strip()).first()
                    if uom:
                        item.uom_id = uom.id
                
                # Text fields
                if row_data.get('Supplier Name'):
                    item.supplier_name = str(row_data['Supplier Name']).strip()
                
                # Numeric fields
                try:
                    if row_data.get('Min Level'):
                        item.min_level = float(row_data['Min Level'])
                    if row_data.get('Max Level'):
                        item.max_level = float(row_data['Max Level'])
                    if row_data.get('Price Per Kg'):
                        item.price_per_kg = float(row_data['Price Per Kg'])
                    if row_data.get('Price Per UOM'):
                        item.price_per_uom = float(row_data['Price Per UOM'])
                    if row_data.get('Kg Per Unit'):
                        item.kg_per_unit = float(row_data['Kg Per Unit'])
                    if row_data.get('Units Per Bag'):
                        item.units_per_bag = int(row_data['Units Per Bag'])
                    if row_data.get('Loss Percentage'):
                        item.loss_percentage = float(row_data['Loss Percentage'])
                except (ValueError, TypeError):
                    # If conversion fails, skip the numeric field
                    pass
                
                # Boolean fields
                if row_data.get('Is Make To Order'):
                    value = str(row_data['Is Make To Order']).strip().lower()
                    item.is_make_to_order = value in ['true', '1', 'yes', 'y']
                
                if row_data.get('Is Active'):
                    value = str(row_data['Is Active']).strip().lower()
                    item.is_active = value in ['true', '1', 'yes', 'y']
                else:
                    item.is_active = True  # Default to active
                
                db.session.add(item)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f'Row {row_num}: {str(e)}')
        
        db.session.commit()
        
        message = f'Upload completed: {success_count} items added successfully'
        if error_count > 0:
            message += f', {error_count} errors occurred'
            if len(errors) <= 5:  # Show first 5 errors
                message += f'. Errors: {"; ".join(errors)}'
        
        return jsonify({'message': message}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process Excel file: {str(e)}'}), 500

@item_master_bp.route('/item-master/download-excel', methods=['GET'])
def download_excel():
    try:
        # Get search parameters
        search_code = request.args.get('item_code', '').strip()
        search_description = request.args.get('description', '').strip()
        search_type = request.args.get('item_type', '').strip()
        
        # Build query with same logic as get_items
        query = ItemMaster.query
        
        if search_code:
            query = query.filter(ItemMaster.item_code.ilike(f"%{search_code}%"))
        if search_description:
            query = query.filter(ItemMaster.description.ilike(f"%{search_description}%"))
        if search_type:
            item_type_obj = ItemType.query.filter_by(type_name=search_type).first()
            if item_type_obj:
                query = query.filter(ItemMaster.item_type_id == item_type_obj.id)
        
        items = query.all()
        
        # Create workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Item Master"
        
        # Define headers
        headers = [
            'Item Code', 'Description', 'Type', 'Category', 'Department', 
            'UOM', 'Min Level', 'Max Level', 'Price Per Kg', 'Price Per UOM', 
            'Supplier Name', 'Is Make To Order', 'Kg Per Unit', 'Units Per Bag', 
            'Loss Percentage', 'Is Active'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row, item in enumerate(items, 2):
            sheet.cell(row=row, column=1, value=item.item_code)
            sheet.cell(row=row, column=2, value=item.description)
            sheet.cell(row=row, column=3, value=item.item_type.type_name if item.item_type else '')
            sheet.cell(row=row, column=4, value=item.category.name if item.category else '')
            sheet.cell(row=row, column=5, value=item.department.departmentName if item.department else '')
            sheet.cell(row=row, column=6, value=item.uom.UOMName if item.uom else '')
            sheet.cell(row=row, column=7, value=item.min_level)
            sheet.cell(row=row, column=8, value=item.max_level)
            sheet.cell(row=row, column=9, value=item.price_per_kg)
            sheet.cell(row=row, column=10, value=item.price_per_uom)
            sheet.cell(row=row, column=11, value=item.supplier_name or '')
            sheet.cell(row=row, column=12, value='Yes' if item.is_make_to_order else 'No')
            sheet.cell(row=row, column=13, value=item.kg_per_unit)
            sheet.cell(row=row, column=14, value=item.units_per_bag)
            sheet.cell(row=row, column=15, value=item.loss_percentage)
            sheet.cell(row=row, column=16, value='Yes' if item.is_active else 'No')
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f'item_master_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate Excel file: {str(e)}'}), 500

@item_master_bp.route('/item-master/download-template', methods=['GET'])
def download_template():
    try:
        # Create workbook with template structure
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Item Master Template"
        
        # Define headers with descriptions
        headers = [
            'Item Code', 'Description', 'Type', 'Category', 'Department', 
            'UOM', 'Min Level', 'Max Level', 'Price Per Kg', 'Price Per UOM', 
            'Supplier Name', 'Is Make To Order', 'Kg Per Unit', 'Units Per Bag', 
            'Loss Percentage', 'Is Active'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add sample data row
        sample_data = [
            'ITEM001', 'Sample Item Description', 'Raw Material', 'Category Name', 'Department Name',
            'KG', '10', '100', '5.50', '6.00', 'ABC Suppliers Ltd', 'No', '', '', '', 'Yes'
        ]
        
        for col, value in enumerate(sample_data, 1):
            sheet.cell(row=2, column=col, value=value)
        
        # Add instructions in a separate sheet
        instructions_sheet = workbook.create_sheet("Instructions")
        instructions = [
            "ITEM MASTER UPLOAD INSTRUCTIONS",
            "",
            "Required Columns:",
            "- Item Code: Unique identifier for the item",
            "- Description: Item description",
            "- Type: Must match existing item types in the system",
            "",
            "Optional Columns:",
            "- Category: Must match existing categories",
            "- Department: Must match existing departments",
            "- UOM: Unit of measure (must match existing UOMs)",
            "- Min Level: Minimum stock level (numeric)",
            "- Max Level: Maximum stock level (numeric)",
            "- Price Per Kg: Price per kilogram (numeric, for Raw Materials)",
            "- Price Per UOM: Price per unit of measure (numeric)",
            "- Supplier Name: Name of the supplier for this item",
            "- Is Make To Order: Yes/No (for Finished Goods)",
            "- Kg Per Unit: Kilograms per unit (numeric, for Finished Goods)",
            "- Units Per Bag: Units per bag (numeric, for Finished Goods)",
            "- Loss Percentage: Loss percentage (numeric, for Finished Goods)",
            "- Is Active: Yes/No (defaults to Yes if not specified)",
            "",
            "Notes:",
            "- Do not modify the header row",
            "- Empty rows will be skipped",
            "- Items with duplicate codes will be rejected",
            "- Invalid references (category, department, etc.) will be ignored",
            "- Boolean fields accept: Yes/No, True/False, 1/0, Y/N"
        ]
        
        for row, instruction in enumerate(instructions, 1):
            cell = instructions_sheet.cell(row=row, column=1, value=instruction)
            if row == 1:  # Title
                cell.font = Font(bold=True, size=14)
            elif instruction.endswith(":"):  # Section headers
                cell.font = Font(bold=True)
        
        # Auto-adjust column widths for both sheets
        for sheet_obj in [sheet, instructions_sheet]:
            for column in sheet_obj.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                sheet_obj.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name='item_master_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate template: {str(e)}'}), 500

@item_master_bp.route('/item-master/test-form', methods=['GET'])
def test_form():
    """Serve a simple test form for debugging form submission issues."""
    import os
    test_file_path = os.path.join(os.getcwd(), 'test_form_simple.html')
    if os.path.exists(test_file_path):
        with open(test_file_path, 'r', encoding='utf-8') as f:
            return f.read()
    else:
        return "Test form file not found", 404

@item_master_bp.route('/item-master/upload', methods=['GET', 'POST'])
def item_master_upload():
    """Render the upload page for Item Master Excel/CSV upload and handle file uploads."""
    if request.method == 'POST':
        # Handle file upload - reuse the existing upload_excel logic
        try:
            if 'file' not in request.files:
                flash('No file uploaded', 'error')
                return render_template('item_master/upload.html', current_page='item_master')
            
            file = request.files['file']
            if file.filename == '':
                flash('No file selected', 'error')
                return render_template('item_master/upload.html', current_page='item_master')
            
            if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
                flash('Please upload a valid Excel or CSV file (.xlsx, .xls, .csv)', 'error')
                return render_template('item_master/upload.html', current_page='item_master')
            
            # Process the file using existing upload logic
            # Read the Excel file
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            
            # Expected columns: Item Code, Description, Type, Category, Department, UOM, Min Level, Max Level, Price Per Kg, Is Make To Order, Kg Per Unit, Units Per Bag, Loss Percentage, Is Active
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Validate required headers
            required_headers = ['Item Code', 'Description', 'Type']
            missing_headers = [h for h in required_headers if h not in headers]
            if missing_headers:
                flash(f'Missing required columns: {", ".join(missing_headers)}', 'error')
                return render_template('item_master/upload.html', current_page='item_master')
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Process each row (skip header)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not any(row):  # Skip empty rows
                        continue
                    
                    # Create dictionary from row data
                    row_data = dict(zip(headers, row))
                    
                    # Skip if required fields are empty
                    if not row_data.get('Item Code') or not row_data.get('Description') or not row_data.get('Type'):
                        error_count += 1
                        errors.append(f'Row {row_num}: Missing required fields')
                        continue
                    
                    item_code = str(row_data['Item Code']).strip()
                    description = str(row_data['Description']).strip()
                    item_type = str(row_data['Type']).strip()
                    
                    # Check if item already exists
                    existing_item = ItemMaster.query.filter_by(item_code=item_code).first()
                    if existing_item:
                        error_count += 1
                        errors.append(f'Row {row_num}: Item code {item_code} already exists')
                        continue
                    
                    # Validate item type
                    valid_item_type = ItemType.query.filter_by(type_name=item_type).first()
                    if not valid_item_type:
                        error_count += 1
                        errors.append(f'Row {row_num}: Invalid item type "{item_type}"')
                        continue
                    
                    # Create new item
                    item = ItemMaster()
                    item.item_code = item_code
                    item.description = description
                    # Find ItemType by name and set the foreign key
                    valid_item_type = ItemType.query.filter_by(type_name=item_type).first()
                    if valid_item_type:
                        item.item_type_id = valid_item_type.id
                    else:
                        # Skip this row if item type is invalid
                        error_count += 1
                        continue
                    
                    # Optional fields
                    if row_data.get('Category'):
                        category = Category.query.filter_by(name=str(row_data['Category']).strip()).first()
                        if category:
                            item.category_id = category.id
                    
                    if row_data.get('Department'):
                        department = Department.query.filter_by(departmentName=str(row_data['Department']).strip()).first()
                        if department:
                            item.department_id = department.id
                    
                    if row_data.get('UOM'):
                        uom = UOM.query.filter_by(UOMName=str(row_data['UOM']).strip()).first()
                        if uom:
                            item.uom_id = uom.id
                    
                    # Text fields
                    if row_data.get('Supplier Name'):
                        item.supplier_name = str(row_data['Supplier Name']).strip()
                    
                    # Numeric fields
                    try:
                        if row_data.get('Min Level'):
                            item.min_level = float(row_data['Min Level'])
                        if row_data.get('Max Level'):
                            item.max_level = float(row_data['Max Level'])
                        if row_data.get('Price Per Kg'):
                            item.price_per_kg = float(row_data['Price Per Kg'])
                        if row_data.get('Price Per UOM'):
                            item.price_per_uom = float(row_data['Price Per UOM'])
                        if row_data.get('Kg Per Unit'):
                            item.kg_per_unit = float(row_data['Kg Per Unit'])
                        if row_data.get('Units Per Bag'):
                            item.units_per_bag = int(row_data['Units Per Bag'])
                        if row_data.get('Loss Percentage'):
                            item.loss_percentage = float(row_data['Loss Percentage'])
                    except (ValueError, TypeError):
                        # If conversion fails, skip the numeric field
                        pass
                    
                    # Boolean fields
                    if row_data.get('Is Make To Order'):
                        value = str(row_data['Is Make To Order']).strip().lower()
                        item.is_make_to_order = value in ['true', '1', 'yes', 'y']
                    
                    if row_data.get('Is Active'):
                        value = str(row_data['Is Active']).strip().lower()
                        item.is_active = value in ['true', '1', 'yes', 'y']
                    else:
                        item.is_active = True  # Default to active
                    
                    db.session.add(item)
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'Row {row_num}: {str(e)}')
            
            db.session.commit()
            
            message = f'Upload completed: {success_count} items added successfully'
            if error_count > 0:
                message += f', {error_count} errors occurred'
                if len(errors) <= 5:  # Show first 5 errors
                    message += f'. Errors: {"; ".join(errors)}'
                flash(message, 'warning')
            else:
                flash(message, 'success')
            
            return render_template('item_master/upload.html', current_page='item_master')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Failed to process file: {str(e)}', 'error')
            return render_template('item_master/upload.html', current_page='item_master')
    
    # GET request - just render the upload page
    return render_template('item_master/upload.html', current_page='item_master')

@item_master_bp.route('/item-master/get-items-by-type/<item_type>', methods=['GET'])
def get_items_by_type(item_type):
    """Get items by item type for hierarchy selection"""
    try:
        # Check if user is authenticated
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': 'Authentication required'}), 401
        
        # Get items by type
        items = ItemMaster.query.join(ItemType).filter(
            ItemType.type_name == item_type,
            ItemMaster.is_active == True
        ).order_by(ItemMaster.item_code).all()
        
        items_data = []
        for item in items:
            items_data.append({
                'id': item.id,
                'item_code': item.item_code,
                'description': item.description or ''
            })
        
        return jsonify({'success': True, 'items': items_data})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@item_master_bp.route('/autocomplete-item-code', methods=['GET'])
def autocomplete_item_code():
    query = request.args.get('query', '').strip()
    if not query:
        return jsonify([])
    
    # Search for items matching the query
    items = ItemMaster.query.filter(
        ItemMaster.item_code.ilike(f"%{query}%")
    ).limit(50).all()
    
    suggestions = []
    for item in items:
        suggestions.append({
            'item_code': item.item_code,
            'description': item.description,
            'item_type': item.item_type.type_name if item.item_type else None
        })
    
    return jsonify(suggestions)

@item_master_bp.route('/search_item_codes', methods=['GET'])
def search_item_codes():
    # Check if user is authenticated
    if 'user_id' not in session:
        return jsonify([]), 401
    
    term = request.args.get('term', '')
    if not term or len(term) < 2:
        return jsonify([])
    
    try:
        # Search for items that match the term in either item_code or description
        from sqlalchemy import or_
        items = ItemMaster.query.filter(
            or_(
                ItemMaster.item_code.ilike(f'%{term}%'),
                ItemMaster.description.ilike(f'%{term}%')
            )
        ).limit(50).all()
        
        # Return list of matching items with descriptions
        results = [{
            'item_code': item.item_code,
            'description': item.description or ''
        } for item in items]
        
        return jsonify(results)
    except Exception as e:
        return jsonify([]), 500

@item_master_bp.route('/item-master/get-item-info/<int:id>', methods=['GET'])
def get_item_info(id):
    try:
        item = ItemMaster.query.get_or_404(id)
        
        return jsonify({
            'success': True,
            'id': item.id,
            'item_code': item.item_code,
            'description': item.description,
            'avg_weight_per_unit': float(item.avg_weight_per_unit) if item.avg_weight_per_unit else 0,
            'min_level': float(item.min_level) if item.min_level else 0,
            'max_level': float(item.max_level) if item.max_level else 0,
            #'calculation_factor': float(item.calculation_factor) if item.calculation_factor else 0,
            'allergens': [{
                'allergens_id': allergen.allergens_id,
                'name': allergen.name
            } for allergen in item.allergens]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500