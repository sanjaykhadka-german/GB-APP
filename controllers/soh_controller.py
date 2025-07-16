import pandas as pd
import os
import io
from flask import Blueprint, jsonify, render_template, request, redirect, url_for, flash, send_file
from sqlalchemy.sql import text
from sqlalchemy import asc, desc
from werkzeug.utils import secure_filename
from datetime import datetime, date
import pytz
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from database import db

from controllers.packing_controller import update_packing_entry, re_aggregate_filling_and_production_for_week
from models.item_master import ItemMaster
from models.packing import Packing
from models.filling import Filling
from models.production import Production
from models.soh import SOH
# from save_report_data import save_report_data # This line is removed from the top

soh_bp = Blueprint('soh', __name__, template_folder='templates')

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def safe_float(value, default=0.0):
    """Convert value to float safely, returning default if conversion fails."""
    try:
        return float(value)
    except (ValueError, TypeError):
        return default

def create_packing_entry_from_soh(fg_code, description, week_commencing, soh_total_units, item):
    """
    Create or update a packing entry when SOH data is uploaded.
    This is a simplified version specifically for SOH uploads.
    It NO LONGER creates downstream entries directly. This is handled by the re-aggregation function.
    """
    from app import db
    from models.packing import Packing
    import math

    if not item:
        raise ValueError(f"Item not found for code {fg_code}")
        
    # Allow packing creation even without department/machinery - can be set later
    if not item.department_id:
        print(f"Warning: No department assigned for item {fg_code}")
        
    if not item.machinery_id:
        print(f"Warning: No machinery assigned for item {fg_code}")
    
    avg_weight_per_unit = item.kg_per_unit or item.avg_weight_per_unit or 0.0
    min_level = item.min_level or 0.0
    max_level = item.max_level or 0.0
    calculation_factor = item.calculation_factor or 1.0
    
    # Calculate SOH requirement: how many units we need to reach max_level
    soh_requirement_units_week = max(0, int(max_level - soh_total_units))
    
    # Calculate the requirement in KG
    soh_requirement_kg_week = int(soh_requirement_units_week * avg_weight_per_unit) if avg_weight_per_unit else 0
    
    # Current SOH in KG
    soh_kg = round(soh_total_units * avg_weight_per_unit, 0) if avg_weight_per_unit else 0
    
    # Requirement calculations for packing
    requirement_kg = soh_requirement_kg_week  # This is what we need to pack
    requirement_unit = soh_requirement_units_week  # This is how many units we need to pack
    
    # Total stock calculations (includes what we have + what we'll pack)
    total_stock_kg = soh_kg + requirement_kg
    total_stock_units = soh_total_units + requirement_unit

    existing_packing = Packing.query.filter_by(
        item_id=item.id,
        week_commencing=week_commencing,
        packing_date=week_commencing
    ).first()
    
    if existing_packing:
        packing = existing_packing
        packing.soh_requirement_units_week = soh_requirement_units_week
        packing.soh_units = soh_total_units
        packing.soh_kg = soh_kg
        packing.avg_weight_per_unit = avg_weight_per_unit
        packing.calculation_factor = calculation_factor
        packing.requirement_kg = requirement_kg
        packing.requirement_unit = requirement_unit
        packing.department_id = item.department_id
        packing.machinery_id = item.machinery_id
        packing.total_stock_kg = total_stock_kg
        packing.total_stock_units = total_stock_units
    else:
        packing = Packing(
            packing_date=week_commencing,
            item_id=item.id,
            week_commencing=week_commencing,
            soh_requirement_kg_week=soh_requirement_kg_week,
            soh_requirement_units_week=soh_requirement_units_week,
            soh_units=soh_total_units,
            soh_kg=soh_kg,
            avg_weight_per_unit=avg_weight_per_unit,
            calculation_factor=calculation_factor,
            requirement_kg=requirement_kg,
            requirement_unit=requirement_unit,
            department_id=item.department_id,
            machinery_id=item.machinery_id,
            total_stock_kg=total_stock_kg,
            total_stock_units=total_stock_units
        )
        db.session.add(packing)
        
    # Commit the packing entry to database
    db.session.commit()
        
    return True, "Packing entry queued for creation."

@soh_bp.route('/soh_upload', methods=['GET', 'POST'])
def soh_upload():
    if request.method == 'POST':
        try:
            if 'file' not in request.files:
                flash('No file uploaded', 'danger')
                return render_template('soh/upload.html', current_page="soh")

            file = request.files['file']
            sheet_name = request.form.get('sheet_name', '').strip() or 'soh_table1'  # Default to soh_table1
            form_week_commencing = request.form.get('week_commencing', '').strip()
            
            print(f"Form data - sheet_name: '{sheet_name}', week_commencing: '{form_week_commencing}'")

            if file.filename == '':
                flash('No file selected', 'danger')
                return render_template('soh/upload.html', current_page="soh")

            if not allowed_file(file.filename):
                flash('Invalid file type', 'danger')
                return render_template('soh/upload.html', current_page="soh")

            # Read Excel file
            df = pd.read_excel(file, sheet_name=sheet_name)
            
            # Debug: Print column names and first few rows
            print("\nExcel columns:", df.columns.tolist())
            print("\nFirst few rows:")
            print(df.head())
            
            # Clean up column names - strip whitespace and handle case
            df.columns = [str(col).strip() for col in df.columns]
            
            # Map expected column names to possible variations
            column_map = {
                'FG Code': ['FG Code', 'FGCode', 'FG_Code', 'fg_code', 'Item Code'],
                'Description': ['Description', 'Desc', 'Item Description', 'description'],
                'Soh_total_Box': ['Soh_total_Box', 'SOH Total Boxes', 'Total Boxes', 'soh_total_boxes'],
                'Soh_total_Unit': ['Soh_total_Unit', 'SOH Total Units', 'Total Units', 'soh_total_units'],
                'Week Commencing': ['Week Commencing', 'week_commencing', 'Week_Commencing']
            }
            
            # Find actual column names in the Excel file
            actual_columns = {}
            for expected_col, variations in column_map.items():
                found = False
                for var in variations:
                    if var in df.columns:
                        actual_columns[expected_col] = var
                        found = True
                        break
                # Week Commencing is optional in the file
                if not found and expected_col != 'Week Commencing':
                    flash(f'Required column "{expected_col}" not found. Looked for variations: {variations}', 'danger')
                    return render_template('soh/upload.html', current_page="soh")
            
            print("\nMapped columns:", actual_columns)
            
            # Clean up data - replace NaN with empty strings for text fields and 0 for numeric fields
            df[actual_columns['FG Code']] = df[actual_columns['FG Code']].fillna('').astype(str)
            df[actual_columns['Description']] = df[actual_columns['Description']].fillna('').astype(str)
            df[actual_columns['Soh_total_Box']] = df[actual_columns['Soh_total_Box']].fillna(0)
            df[actual_columns['Soh_total_Unit']] = df[actual_columns['Soh_total_Unit']].fillna(0)
            
            # Track success/errors
            success_count = 0
            error_count = 0
            errors = []
            
            # Track all week commencing dates that are affected by this upload
            affected_weeks = set()

            for index, row in df.iterrows():
                fg_code = str(row[actual_columns['FG Code']]).strip()
                if not fg_code:
                    continue  # Skip empty rows

                # Determine the week_commencing for the current row
                row_week_commencing_str = str(row.get(actual_columns.get('Week Commencing'), '')).strip()
                
                if form_week_commencing:
                    week_commencing = datetime.strptime(form_week_commencing, '%Y-%m-%d').date()
                elif row_week_commencing_str:
                    try:
                        # Handle different date formats from Excel
                        week_commencing = pd.to_datetime(row_week_commencing_str).date()
                    except Exception:
                        error_msg = f"Invalid 'Week Commencing' date format in row {index + 2}: {row_week_commencing_str}"
                        errors.append(error_msg)
                        error_count += 1
                        continue
                else:
                    # Fallback if no date is provided at all
                    error_msg = f"Missing 'Week Commencing' date in row {index + 2} and no date provided in the form."
                    errors.append(error_msg)
                    error_count += 1
                    continue
                    
                # Add the current week to our set of affected weeks
                affected_weeks.add(week_commencing)

                try:
                    description = str(row[actual_columns['Description']]).strip()
                    soh_total_boxes = int(safe_float(row[actual_columns['Soh_total_Box']]))
                    soh_total_units = int(safe_float(row[actual_columns['Soh_total_Unit']]))
                    
                    # Find item master record
                    item = ItemMaster.query.filter_by(item_code=fg_code).first()

                    # Create or update SOH entry
                    soh_entry = SOH.query.filter_by(item_id=item.id if item else None, week_commencing=week_commencing).first()
                    if soh_entry:
                        soh_entry.soh_total_boxes = soh_total_boxes
                        soh_entry.soh_total_units = soh_total_units
                        soh_entry.edit_date = datetime.now(pytz.timezone('Australia/Sydney'))
                    elif item:
                        soh_entry = SOH(
                            item_id=item.id,
                            fg_code=fg_code,
                            week_commencing=week_commencing,
                            description=description,
                            soh_total_boxes=soh_total_boxes,
                            soh_total_units=soh_total_units,
                            edit_date=datetime.now(pytz.timezone('Australia/Sydney'))
                        )
                        db.session.add(soh_entry)
                    
                    # Only create packing entries if we have a valid item
                    if item:
                        # Create packing entry - no longer creates downstream here
                        create_packing_entry_from_soh(
                            fg_code=item.item_code,
                            description=item.description,
                            week_commencing=week_commencing,
                            soh_total_units=soh_total_units,
                            item=item
                        )
                        success_count += 1
                    else:
                        # Still count as success if we can save the SOH record
                        success_count += 1
                        print(f"Warning: Item not found in ItemMaster for FG Code {fg_code}. SOH record created without downstream entries.")
                        
                except Exception as e:
                    error_msg = f"Error processing row for FG Code {fg_code}: {str(e)}"
                    print(error_msg)
                    errors.append(error_msg)
                    error_count += 1
                    continue
            
            # Commit transaction if no errors
            if error_count == 0:
                db.session.commit()
                flash(f'Successfully processed {success_count} SOH records.', 'success')

                # After successful commit, run aggregation for all affected weeks
                aggregation_success_count = 0
                aggregation_error_count = 0
                for week in affected_weeks:
                    print(f"Running aggregation for week: {week}")
                    success, message = re_aggregate_filling_and_production_for_week(week)
                    if success:
                        aggregation_success_count += 1
                    else:
                        aggregation_error_count += 1
                        print(f"Aggregation failed for week {week}: {message}")

                if aggregation_error_count > 0:
                    flash(f'⚠️ Downstream aggregation failed for {aggregation_error_count} out of {len(affected_weeks)} week(s). Please check logs.', 'warning')
                else:
                    flash(f'✅ Successfully re-aggregated downstream requirements for {len(affected_weeks)} week(s).', 'success')
            else:
                db.session.rollback()
                flash(f'Processed {success_count} records with {error_count} errors. First 5 errors: {"; ".join(errors[:5])}', 'danger')

        except Exception as e:
            db.session.rollback()
            flash(f'An unexpected error occurred: {str(e)}', 'danger')
            print(f"Error during SOH upload: {str(e)}")

        # After SOH is saved, populate the report tables for the affected weeks
        try:
            from save_report_data import save_report_data
            from populate_inventory import populate_inventory
            
            if affected_weeks:
                print(f"Calling save_report_data for weeks: {affected_weeks}")
                save_report_data(affected_weeks)
                print("Finished calling save_report_data.")
                flash('Report tables generated successfully for the uploaded week(s)!', 'success')
                
                # Now, populate the inventory based on the new reports
                print(f"Calling populate_inventory for weeks: {affected_weeks}")
                populate_inventory(affected_weeks)
                print("Finished calling populate_inventory.")
                flash('Inventory generated successfully for the uploaded week(s)!', 'success')
            else:
                print("No affected weeks to process.")
                flash('SOH data processed, but no weeks were identified for report/inventory generation.', 'info')
        except Exception as e:
            print(f"Error during post-upload processing: {str(e)}")
            flash(f"SOH data uploaded, but failed during post-processing: {str(e)}", 'warning')

        return redirect(url_for('soh.soh_list'))
    
    return render_template('soh/upload.html', current_page="soh")

@soh_bp.route('/download_template')
def download_template():
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'soh_table1'  # Set the sheet name to match what we expect
        
        # Define headers with exact column names needed
        headers = [
            'Week Commencing',
            'FG Code',
            'Description',
            'Soh_dispatch_Box',
            'Soh_dispatch_Unit', 
            'Soh_packing_Box',
            'Soh_packing_Unit',
            'Soh_total_Box',
            'Soh_total_Unit'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Add sample data
        sample_data = [
            '2024-01-01',  # Week Commencing
            'FG001',       # FG Code
            'Sample Item', # Description
            '10',         # soh_dispatch_boxes
            '100',        # soh_dispatch_units
            '20',         # soh_packing_boxes
            '200',        # soh_packing_units
            '30',         # soh_total_boxes
            '300'         # soh_total_units
        ]
        
        # Add sample row
        for col, value in enumerate(sample_data, 1):
            sheet.cell(row=2, column=col, value=value)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='soh_upload_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating template: {str(e)}", "danger")
        return redirect(url_for('soh.soh_upload'))

@soh_bp.route('/soh_list', methods=['GET'])
def soh_list():
    from app import db
    from models.item_master import ItemMaster

    search_fg_code = request.args.get('fg_code', '').strip()
    search_description = request.args.get('description', '').strip()
    search_week_commencing = request.args.get('week_commencing', '').strip() # This will be YYYY-MM-DD from frontend
    sort_by = request.args.get('sort_by', 'id').strip()
    sort_direction = request.args.get('sort_direction', 'asc').strip()

    sohs = [] # Initialize sohs as an empty list
    try:
        sohs_query = SOH.query.join(ItemMaster, SOH.item_id == ItemMaster.id)
        if search_fg_code:
            sohs_query = sohs_query.filter(ItemMaster.item_code.ilike(f"%{search_fg_code}%"))
        if search_description:
            sohs_query = sohs_query.filter(SOH.description.ilike(f"%{search_description}%"))
        if search_week_commencing:
            try:
                # Parse YYYY-MM-DD from the frontend search input
                week_commencing_date = datetime.strptime(search_week_commencing, '%Y-%m-%d').date()
                sohs_query = sohs_query.filter(SOH.week_commencing == week_commencing_date)
            except ValueError as e:
                flash(f"Invalid Week Commencing date format in search: {str(e)}. Please use YYYY-MM-DD.", "danger")
                return render_template('soh/list.html',
                                       sohs=[], # Pass an empty list on error
                                       search_fg_code=search_fg_code,
                                       search_description=search_description,
                                       search_week_commencing=search_week_commencing,
                                       sort_by=sort_by,
                                       sort_direction=sort_direction,
                                       current_page="soh")

        # Apply sorting
        if sort_by in ['week_commencing', 'fg_code', 'description', 'edit_date']:
            if sort_direction == 'desc':
                sohs_query = sohs_query.order_by(desc(getattr(SOH, sort_by)))
            else:
                sohs_query = sohs_query.order_by(asc(getattr(SOH, sort_by)))

        sohs = sohs_query.all()

        for soh in sohs:
            # Format dates for display in DD-MM-YYYY format for the initial render
            soh.week_commencing_str = soh.week_commencing.strftime('%d-%m-%Y') if soh.week_commencing else ''
            soh.week_commencing_input_str = soh.week_commencing.strftime('%Y-%m-%d') if soh.week_commencing else ''
            soh.edit_date_str = soh.edit_date.strftime('%d-%m-%Y %H:%M:%S') if soh.edit_date else ''

    except Exception as e:
        flash(f"Error fetching SOH list: {str(e)}", "danger")
        sohs = [] # Ensure it's an empty list if an exception occurs

    return render_template('soh/list.html',
                           sohs=sohs,
                           search_fg_code=search_fg_code,
                           search_description=search_description,
                           search_week_commencing=search_week_commencing, # Pass YYYY-MM-DD back to keep form value
                           sort_by=sort_by,
                           sort_direction=sort_direction,
                           current_page="soh")

@soh_bp.route('/soh_create', methods=['GET', 'POST'])
def soh_create():
    from app import db
    from models.item_master import ItemMaster

    if request.method == 'POST':
        try:
            fg_code = request.form['fg_code'].strip()
            week_commencing = request.form.get('week_commencing')
            description = request.form['description'].strip()
            # Convert empty strings from form to 0.0 or None as needed
            dispatch_boxes = float(request.form.get('soh_dispatch_boxes') or 0.0)
            dispatch_units = float(request.form.get('soh_dispatch_units') or 0.0)
            packing_boxes = float(request.form.get('soh_packing_boxes') or 0.0)
            packing_units = float(request.form.get('soh_packing_units') or 0.0)

            # Convert week_commencing to date (expected YYYY-MM-DD from HTML date input)
            week_commencing_date = None
            if week_commencing:
                try:
                    week_commencing_date = datetime.strptime(week_commencing, '%Y-%m-%d').date()
                except ValueError as e:
                    flash(f"Invalid Week Commencing date format: {str(e)}. Expected YYYY-MM-DD.", "danger")
                    return redirect(request.url)

            item = ItemMaster.query.filter_by(item_code=fg_code).first()
            if not item:
                return jsonify({"success": False, "error": f"No item found for FG Code: {fg_code}"}), 400
            units_per_bag = item.units_per_bag if item and item.units_per_bag else 1
            avg_weight_per_unit = item.kg_per_unit if item and item.kg_per_unit else 0.0

            soh_total_boxes = dispatch_boxes + packing_boxes
            soh_total_units = (
                (dispatch_boxes * units_per_bag) +
                (packing_boxes * units_per_bag) +
                dispatch_units +
                packing_units
            )

            new_soh = SOH(
                item_id=item.id,  # Use foreign key
                fg_code=fg_code,  # Keep for backward compatibility
                week_commencing=week_commencing_date,
                description=description,
                soh_dispatch_boxes=dispatch_boxes,
                soh_dispatch_units=dispatch_units,
                soh_packing_boxes=packing_boxes,
                soh_packing_units=packing_units,
                soh_total_boxes=soh_total_boxes,
                soh_total_units=soh_total_units,
                edit_date=datetime.now(pytz.timezone('Australia/Sydney'))
            )
            db.session.add(new_soh)
            db.session.commit()

            if soh_total_boxes >= 0 or soh_total_units >= 0: # Condition for update_packing_entry
                # Ensure packing_date is a date object
                packing_date_for_update = week_commencing_date if week_commencing_date else date.today()
                success = create_packing_entry_from_soh(
                    fg_code=fg_code,
                    description=description,
                    week_commencing=packing_date_for_update,
                    soh_total_units=soh_total_units,
                    item=item
                )
                if not success:
                    flash(f"Warning: Could not update Packing for FG Code {fg_code}", "warning")
                
                # Create downstream Filling and Production entries
                try:
                    from controllers.bom_service import BOMService
                    
                    # Create Filling entry for WIPF
                    filling_entry = BOMService.create_filling_entry(
                        item_id=item.id,
                        week_commencing=packing_date_for_update,
                        requirement_kg=0,  # Will be calculated from all packing entries
                        requirement_unit=0
                    )
                    if filling_entry:
                        db.session.commit()
                    
                    # Create Production entry for WIP
                    production_entry = BOMService.create_production_entry(
                        item_id=item.id,
                        week_commencing=packing_date_for_update,
                        requirement_kg=0,  # Will be calculated from all packing entries
                        requirement_unit=0
                    )
                    if production_entry:
                        db.session.commit()
                        
                    # ADD THIS: Create usage reports for the production entry
                    try:
                        usage_reports = BOMService.create_usage_report(
                            item_id=item.id,
                            week_commencing=packing_date_for_update,
                            requirement_kg=production_entry.total_kg,
                            requirement_unit=0
                        )
                        if usage_reports:
                            db.session.commit()
                            print(f"Created {len(usage_reports)} usage report entries for {fg_code}")
                    except Exception as e:
                        print(f"Warning: Could not create usage reports for {fg_code}: {str(e)}")
                        
                except Exception as e:
                    flash(f"Warning: Could not create downstream entries for {fg_code}: {str(e)}", "warning")

            flash("SOH entry created successfully!", "success")
            return redirect(url_for('soh.soh_list'))

        except Exception as e:
            db.session.rollback()
            flash(f"Error creating SOH entry: {str(e)}", "danger")
            return redirect(request.url)

    fg_code = request.args.get('fg_code', '')
    return render_template('soh/create.html', fg_code=fg_code, current_page="soh")

@soh_bp.route('/soh_edit/<int:id>', methods=['GET', 'POST'])
def soh_edit(id):
    from app import db
    from models.item_master import ItemMaster

    soh = SOH.query.get_or_404(id)

    if request.method == 'POST':
        try:
            fg_code = request.form['fg_code'].strip()
            week_commencing = request.form.get('week_commencing')
            description = request.form['description'].strip()
            soh_dispatch_boxes = float(request.form.get('soh_dispatch_boxes') or 0.0)
            soh_dispatch_units = float(request.form.get('soh_dispatch_units') or 0.0)
            soh_packing_boxes = float(request.form.get('soh_packing_boxes') or 0.0)
            soh_packing_units = float(request.form.get('soh_packing_units') or 0.0)

            # Convert week_commencing to date (expected YYYY-MM-DD from HTML date input)
            week_commencing_date = None
            if week_commencing:
                try:
                    week_commencing_date = datetime.strptime(week_commencing, '%Y-%m-%d').date()
                except ValueError:
                    # Fallback for DD-MM-YYYY, although YYYY-MM-DD is expected from input type="date"
                    try:
                        week_commencing_date = datetime.strptime(week_commencing, '%d-%m-%Y').date()
                    except ValueError as e:
                        flash(f"Invalid Week Commencing date format: {str(e)}. Please use YYYY-MM-DD.", "danger")
                        return redirect(request.url)

            item = ItemMaster.query.filter_by(item_code=fg_code).first()
            if not item:
                flash(f"FG Code '{fg_code}' not found in Item Master table.", "danger")
                return redirect(request.url)
            units_per_bag = item.units_per_bag if item and item.units_per_bag else 1
            avg_weight_per_unit = item.kg_per_unit if item and item.kg_per_unit else 0.0

            # Recalculate totals
            soh_total_boxes = soh_dispatch_boxes + soh_packing_boxes
            soh_total_units = (
                (soh_dispatch_boxes * units_per_bag) +
                (soh_packing_boxes * units_per_bag) +
                soh_dispatch_units +
                soh_packing_units
            )

            # Update item_id if fg_code changed
            if fg_code != (soh.item.item_code if soh.item else soh.fg_code):
                new_item = ItemMaster.query.filter_by(item_code=fg_code).first()
                if new_item:
                    soh.item_id = new_item.id
                else:
                    flash(f"No item found for FG Code: {fg_code}", "danger")
                    return redirect(request.url)
            soh.fg_code = fg_code  # Keep for backward compatibility
            soh.week_commencing = week_commencing_date
            soh.description = description
            soh.soh_dispatch_boxes = soh_dispatch_boxes
            soh.soh_dispatch_units = soh_dispatch_units
            soh.soh_packing_boxes = soh_packing_boxes
            soh.soh_packing_units = soh_packing_units
            soh.soh_total_boxes = soh_total_boxes # Use calculated values
            soh.soh_total_units = soh_total_units # Use calculated values
            soh.edit_date = datetime.now(pytz.timezone('Australia/Sydney'))

            db.session.commit()

            if soh_total_boxes > 0 or soh_total_units > 0: # Condition for update_packing_entry
                # Fetch item to get calculation_factor
                item = ItemMaster.query.filter_by(item_code=fg_code).first()
                avg_weight_per_unit = item.kg_per_unit if item and item.kg_per_unit else 0.0
                
                packing_date = week_commencing_date if week_commencing_date else date.today()
                success = create_packing_entry_from_soh(
                    fg_code=fg_code,
                    description=description,
                    week_commencing=packing_date,
                    soh_total_units=soh_total_units,
                    item=item
                )
                if not success:
                    flash(f"Warning: Could not update Packing for FG Code {fg_code}", "warning")
                
                # Create downstream Filling and Production entries
                try:
                    from controllers.bom_service import BOMService
                    
                    # Create Filling entry for WIPF
                    filling_entry = BOMService.create_filling_entry(
                        item_id=item.id,
                        week_commencing=packing_date,
                        requirement_kg=0,  # Will be calculated from all packing entries
                        requirement_unit=0
                    )
                    if filling_entry:
                        db.session.commit()
                    
                    # Create Production entry for WIP
                    production_entry = BOMService.create_production_entry(
                        item_id=item.id,
                        week_commencing=packing_date,
                        requirement_kg=0,  # Will be calculated from all packing entries
                        requirement_unit=0
                    )
                    if production_entry:
                        db.session.commit()
                        
                except Exception as e:
                    flash(f"Warning: Could not create downstream entries for {fg_code}: {str(e)}", "warning")

            flash("SOH entry updated successfully!", "success")
            return redirect(url_for('soh.soh_list'))

        except Exception as e:
            db.session.rollback()
            flash(f"Error updating SOH entry: {str(e)}", "danger")
            return redirect(request.url)

    # Format edit_date and week_commencing for display in the form (YYYY-MM-DD for input type="date")
    soh.edit_date_str = soh.edit_date.strftime('%d-%m-%Y %H:%M:%S') if soh.edit_date else ''
    soh.week_commencing_str = soh.week_commencing.strftime('%Y-%m-%d') if soh.week_commencing else '' # For <input type="date">

    return render_template('soh/edit.html', soh=soh, current_page="soh")


@soh_bp.route('/soh_delete/<int:id>', methods=['POST'])
def soh_delete(id):
    from app import db
    from models.soh import SOH
    # from datetime import datetime # Already imported

    try:
        soh = SOH.query.get_or_404(id)
        db.session.delete(soh)
        db.session.commit()
        flash("SOH entry deleted successfully!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting SOH entry: {str(e)}", "danger")

    return redirect(url_for('soh.soh_list'))

@soh_bp.route('/autocomplete_soh', methods=['GET'])
def autocomplete_soh():
    from app import db
    from models.item_master import ItemMaster

    search = request.args.get('query', '').strip()

    if not search:
        return jsonify([])

    try:
        # Use SQLAlchemy's ORM for better integration and less raw SQL
        results = db.session.query(ItemMaster.item_code, ItemMaster.description).join(
            ItemMaster.item_type
        ).filter(
            ItemMaster.item_code.ilike(f"{search}%"),
            ItemMaster.item_type.has(type_name='FG') | ItemMaster.item_type.has(type_name='WIPF')
        ).limit(10).all()
        suggestions = [{"fg_code": row.item_code, "description": row.description} for row in results]
        return jsonify(suggestions)
    except Exception as e:
        print("Error fetching SOH autocomplete suggestions:", e)
        return jsonify([])


@soh_bp.route('/get_search_sohs', methods=['GET'])
def get_search_sohs():
    from app import db
    from models.soh import SOH
    from models.item_master import ItemMaster
    from datetime import datetime
    from sqlalchemy import desc, asc
    import traceback

    try:
        print("Starting get_search_sohs function")
        search_fg_code = request.args.get('fg_code', '').strip()
        search_description = request.args.get('description', '').strip()
        search_week_commencing = request.args.get('week_commencing', '').strip()
        sort_by = request.args.get('sort_by', 'id').strip()
        sort_direction = request.args.get('sort_direction', 'asc').strip()

        print(f"Search parameters - FG Code: {search_fg_code}, Description: {search_description}, Week Commencing: {search_week_commencing}")
        print(f"Sort parameters - Sort By: {sort_by}, Direction: {sort_direction}")

        # Start with base query - use LEFT OUTER JOIN with ItemMaster
        sohs_query = db.session.query(SOH).outerjoin(ItemMaster, SOH.item_id == ItemMaster.id)

        # Apply filters
        if search_fg_code:
            sohs_query = sohs_query.filter(
                db.or_(
                    ItemMaster.item_code.ilike(f"%{search_fg_code}%"),
                    SOH.fg_code.ilike(f"%{search_fg_code}%")
                )
            )
        if search_description:
            sohs_query = sohs_query.filter(SOH.description.ilike(f"%{search_description}%"))
        if search_week_commencing:
            try:
                # Expect YYYY-MM-DD from frontend search input
                week_commencing_date = datetime.strptime(search_week_commencing, '%Y-%m-%d').date()
                sohs_query = sohs_query.filter(SOH.week_commencing == week_commencing_date)
            except ValueError as e:
                print(f"Invalid date format: {e}")
                return jsonify({"error": f"Invalid date format: {str(e)}"}), 400

        # Apply sorting
        if sort_by == 'fg_code':
            sohs_query = sohs_query.order_by(
                desc(db.func.coalesce(ItemMaster.item_code, SOH.fg_code)) if sort_direction == 'desc' 
                else asc(db.func.coalesce(ItemMaster.item_code, SOH.fg_code))
            )
        elif sort_by in ['week_commencing', 'description', 'edit_date']:
            sohs_query = sohs_query.order_by(
                desc(getattr(SOH, sort_by)) if sort_direction == 'desc' 
                else asc(getattr(SOH, sort_by))
            )
        else:
            # Default sort by id
            sohs_query = sohs_query.order_by(
                desc(SOH.id) if sort_direction == 'desc' else asc(SOH.id)
            )

        # Execute query and print SQL for debugging
        print(f"SQL Query: {sohs_query}")
        results = sohs_query.all()
        print(f"Found {len(results)} SOH records")

        sohs_data = []
        for soh in results:
            print(f"Processing SOH ID: {soh.id}, Item ID: {soh.item_id}, FG Code: {soh.item.item_code if soh.item else soh.fg_code}")

            # Format dates for display
            week_commencing_display = soh.week_commencing.strftime('%d-%m-%Y') if soh.week_commencing else ""
            week_commencing_input = soh.week_commencing.strftime('%Y-%m-%d') if soh.week_commencing else ""

            # Build response data
            sohs_data.append({
                "id": soh.id,
                "week_commencing": week_commencing_display,  # For displaying in the table (DD-MM-YYYY)
                "week_commencing_original": week_commencing_input,  # For data-original-input attribute (YYYY-MM-DD)
                "fg_code": soh.item.item_code if soh.item else soh.fg_code,  # Use item_code from ItemMaster if available
                "description": soh.description or "",
                "soh_dispatch_boxes": float(soh.soh_dispatch_boxes) if soh.soh_dispatch_boxes is not None else 0.0,
                "soh_dispatch_units": float(soh.soh_dispatch_units) if soh.soh_dispatch_units is not None else 0.0,
                "soh_packing_boxes": float(soh.soh_packing_boxes) if soh.soh_packing_boxes is not None else 0.0,
                "soh_packing_units": float(soh.soh_packing_units) if soh.soh_packing_units is not None else 0.0,
                "soh_total_boxes": float(soh.soh_total_boxes) if soh.soh_total_boxes is not None else 0.0,
                "soh_total_units": float(soh.soh_total_units) if soh.soh_total_units is not None else 0.0,
                "edit_date": soh.edit_date.strftime('%d-%m-%Y %H:%M:%S') if soh.edit_date else ""
            })

        print(f"Returning {len(sohs_data)} records")
        return jsonify(sohs_data)

    except Exception as e:
        error_msg = f"Error fetching search SOHs: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)
        return jsonify({"error": error_msg}), 500

@soh_bp.route('/soh_bulk_edit', methods=['POST'])
def soh_bulk_edit():
    from app import db
    from models.soh import SOH
    from models.item_master import ItemMaster
    # from datetime import datetime # Already imported

    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({"success": False, "error": "No SOH entries selected"}), 400

        week_commencing_str = data.get('week_commencing', '').strip()
        soh_dispatch_boxes_str = data.get('soh_dispatch_boxes', '')
        soh_dispatch_units_str = data.get('soh_dispatch_units', '')
        soh_packing_boxes_str = data.get('soh_packing_boxes', '')
        soh_packing_units_str = data.get('soh_packing_units', '')

        # Convert and validate inputs
        week_commencing_date = None
        if week_commencing_str:
            try:
                week_commencing_date = datetime.strptime(week_commencing_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({"success": False, "error": "Invalid Week Commencing date format. Use YYYY-MM-DD"}), 400

        # Convert numeric fields, allowing empty strings to skip updates (set to None)
        def parse_float_or_none(value_str):
            try:
                return float(value_str) if value_str.strip() != '' else None
            except (ValueError, TypeError):
                return None # Return None if conversion fails

        dispatch_boxes_input = parse_float_or_none(soh_dispatch_boxes_str)
        dispatch_units_input = parse_float_or_none(soh_dispatch_units_str)
        packing_boxes_input = parse_float_or_none(soh_packing_boxes_str)
        packing_units_input = parse_float_or_none(soh_packing_units_str)


        for soh_id in ids:
            soh = SOH.query.get(soh_id) # Use .get() as get_or_404 would raise an error and break the loop
            if not soh:
                print(f"SOH entry with ID {soh_id} not found, skipping.")
                continue # Skip to the next ID if not found

            # Update only fields that were provided (not None)
            if week_commencing_date is not None:
                soh.week_commencing = week_commencing_date
            if dispatch_boxes_input is not None:
                soh.soh_dispatch_boxes = dispatch_boxes_input
            if dispatch_units_input is not None:
                soh.soh_dispatch_units = dispatch_units_input
            if packing_boxes_input is not None:
                soh.soh_packing_boxes = packing_boxes_input
            if packing_units_input is not None:
                soh.soh_packing_units = packing_units_input

            # Use the foreign key relationship if available, fallback to fg_code
            item = soh.item if soh.item else ItemMaster.query.filter_by(item_code=soh.fg_code).first()
            if not item:
                fg_code = soh.item.item_code if soh.item else soh.fg_code
                return jsonify({"success": False, "error": f"No item found for FG Code: {fg_code}"}), 400
            units_per_bag = item.units_per_bag if item and item.units_per_bag else 1
            avg_weight_per_unit = item.kg_per_unit if item and item.kg_per_unit else 0.0


            # Recalculate totals based on potentially updated or existing values
            # Ensure these use actual float/numeric values, not None
            current_dispatch_boxes = soh.soh_dispatch_boxes or 0.0
            current_dispatch_units = soh.soh_dispatch_units or 0.0
            current_packing_boxes = soh.soh_packing_boxes or 0.0
            current_packing_units = soh.soh_packing_units or 0.0

            soh.soh_total_boxes = current_dispatch_boxes + current_packing_boxes
            soh.soh_total_units = (
                (current_dispatch_boxes * units_per_bag) +
                (current_packing_boxes * units_per_bag) +
                current_dispatch_units +
                current_packing_units
            )

            soh.edit_date = datetime.now(pytz.timezone('Australia/Sydney'))

            # Update packing entry if necessary
            # Pass the current soh.week_commencing (which might have just been updated)
            if soh.soh_total_boxes >= 0 or soh.soh_total_units >= 0: # Condition for create_packing_entry_from_soh
                packing_date_for_update = soh.week_commencing if soh.week_commencing else date.today()
                fg_code = soh.item.item_code if soh.item else soh.fg_code
                success = create_packing_entry_from_soh(
                    fg_code=fg_code,
                    description=soh.description,
                    week_commencing=packing_date_for_update,
                    soh_total_units=soh.soh_total_units or 0,
                    item=item
                )
                if not success:
                    print(f"Failed to update Packing for {soh.fg_code} during bulk edit: {success}")

        db.session.commit()
        return jsonify({"success": True})

    except Exception as e:
        db.session.rollback()
        print(f"Error in bulk edit: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

@soh_bp.route('/soh_inline_edit', methods=['POST'])
def soh_inline_edit():
    from app import db
    from models.soh import SOH
    from models.item_master import ItemMaster
    # from datetime import datetime, date # Already imported

    try:
        data = request.get_json()
        print(f"Received data for inline edit: {data}")
        soh_id = data.get('id')
        if not soh_id:
            return jsonify({"success": False, "error": "No SOH ID provided"}), 400

        soh = SOH.query.get_or_404(soh_id)
        field = data.get('field')
        value = data.get(field) # This value is what the frontend sent for the specific field

        if not field:
            return jsonify({"success": False, "error": "No field provided"}), 400

        print(f"Processing field: {field}, value: '{value}'")

        if field == 'week_commencing':
            try:
                if value:
                    # Frontend sends YYYY-MM-DD from the date input
                    parsed_date = datetime.strptime(value, '%Y-%m-%d').date()
                else:
                    parsed_date = None # Allow setting to null/empty
                soh.week_commencing = parsed_date
            except ValueError as e:
                print(f"Date parsing error: {str(e)}")
                return jsonify({"success": False, "error": f"Invalid date format: {str(e)}. Use YYYY-MM-DD."}), 400
        elif field in ['soh_dispatch_boxes', 'soh_dispatch_units', 'soh_packing_boxes', 'soh_packing_units']:
            try:
                # Convert empty string from frontend to 0.0 or None as per database column allows nulls or defaults
                # Assuming these should be 0.0 if empty for calculations.
                numeric_value = float(value) if value is not None and value != '' else 0.0
                setattr(soh, field, numeric_value)
            except (ValueError, TypeError) as e:
                print(f"Number conversion error for {field}: {str(e)}")
                return jsonify({"success": False, "error": f"Invalid number for {field}: {str(e)}."}), 400
        else:
            return jsonify({"success": False, "error": f"Invalid field specified: {field}."}), 400

        # Retrieve current item for calculating totals and updating packing
        item = soh.item if soh.item else ItemMaster.query.filter_by(item_code=soh.fg_code).first()
        if not item:
            fg_code = soh.item.item_code if soh.item else soh.fg_code
            return jsonify({"success": False, "error": f"No item found for FG Code: {fg_code}"}), 400
        units_per_bag = item.units_per_bag if item and item.units_per_bag else 1
        avg_weight_per_unit = item.kg_per_unit if item and item.kg_per_unit else 0.0

        # Recalculate total boxes and units based on potentially updated and existing values
        # Ensure these are treated as 0.0 if they are None for calculation purposes
        current_dispatch_boxes = soh.soh_dispatch_boxes or 0.0
        current_dispatch_units = soh.soh_dispatch_units or 0.0
        current_packing_boxes = soh.soh_packing_boxes or 0.0
        current_packing_units = soh.soh_packing_units or 0.0

        soh.soh_total_boxes = current_dispatch_boxes + current_packing_boxes
        soh.soh_total_units = (
            (current_dispatch_boxes * units_per_bag) +
            (current_packing_boxes * units_per_bag) +
            current_dispatch_units +
            current_packing_units
        )

        soh.edit_date = datetime.now(pytz.timezone('Australia/Sydney'))

        # Update packing entry if necessary
        if soh.soh_total_boxes >= 0 or soh.soh_total_units >= 0:
            # Ensure packing_date is a date object for create_packing_entry_from_soh
            packing_date = soh.week_commencing if soh.week_commencing else date.today()
            success = create_packing_entry_from_soh(
                fg_code=soh.fg_code,
                description=soh.description,
                week_commencing=packing_date,
                soh_total_units=soh.soh_total_units or 0,
                item=item
            )
            if not success:
                print(f"Failed to update Packing for {soh.fg_code} during inline edit: {success}")
                # Consider if you want to flash a warning for inline edits too
                # flash(message, "warning")

        db.session.commit()
        print(f"SOH updated successfully: {soh.id}")
        return jsonify({"success": True})

    except Exception as e:
        db.session.rollback()
        print(f"Error in inline edit: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500