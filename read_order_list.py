import pandas as pd
import numpy as np
from openpyxl import load_workbook

try:
    # Try reading with openpyxl engine
    print("Attempting to read the file...")
    df = pd.read_excel('WC_23062025_18062025 TINO WORKING FILE.xlsm', 
                      sheet_name='Order List',
                      engine='openpyxl')
    
    print("\nSuccessfully read the file!")
    print("Raw data shape:", df.shape)
    print("\nColumns:", list(df.columns))
    print("\nFirst few rows of data:")
    print(df.head(10).to_string())
    
except Exception as e:
    print(f"\nError reading file: {str(e)}")
    
    # Try alternative approach using openpyxl directly
    try:
        print("\nTrying alternative approach...")
        wb = load_workbook('WC_23062025_18062025 TINO WORKING FILE.xlsm', read_only=True, data_only=True)
        print("\nWorkbook sheets:", wb.sheetnames)
        
        if 'Order List' in wb.sheetnames:
            ws = wb['Order List']
            print("\nAccessed 'Order List' sheet")
            print("Sheet dimensions:", ws.dimensions)
            
            # Print first few rows
            print("\nFirst few rows:")
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5)):
                print(f"Row {idx + 1}:", [cell.value for cell in row])
                
    except Exception as e2:
        print(f"\nError with alternative approach: {str(e2)}") 